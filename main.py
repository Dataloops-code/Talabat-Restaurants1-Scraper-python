import asyncio
import json
import os
import tempfile
import sys
import subprocess
from retry import retry
import re
from typing import Dict, List, Tuple
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright
from talabat_main_scraper import TalabatScraper
from SavingOnDrive import SavingOnDrive
from time import sleep
from datetime import datetime
import logging

logging.basicConfig(
    filename='scraper.log',
    level=logging.INFO,  # Changed from DEBUG to INFO
    format='%(asctime)s - %(levelname)s - %(message)s'
)

class MainScraper:
    CURRENT_PROGRESS_FILE = "current_progress.json"
    SCRAPED_PROGRESS_FILE = "scraped_progress.json"

    def __init__(self):
        self.talabat_scraper = TalabatScraper()
        self.output_dir = "output"
        credentials_json = os.environ.get('TALABAT_GCLOUD_KEY_JSON')
        self.drive_uploader = SavingOnDrive(credentials_json=credentials_json)
        
        os.makedirs(self.output_dir, exist_ok=True)
        
        self.current_progress = self.load_current_progress()
        self.scraped_progress = self.load_scraped_progress()
        
        self.github_token = os.environ.get('GITHUB_TOKEN')
        self.ensure_playwright_browsers()

    def ensure_playwright_browsers(self):
        try:
            print("Installing Playwright browsers...")
            subprocess.run([sys.executable, "-m", "playwright", "install", "chromium", "firefox"], 
                          check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            print("Playwright browsers installed successfully")
        except subprocess.CalledProcessError as e:
            print(f"Error installing Playwright browsers: {e}")
            logging.error(f"Error installing Playwright browsers: {e}")

    def load_current_progress(self) -> Dict:
        default_progress = {
            "completed_areas": [],
            "current_area_index": 0,
            "last_updated": None,
            "current_progress": {
                "area_name": None,
                "current_page": 0,
                "total_pages": 0,
                "current_restaurant": 0,
                "total_restaurants": 0,
                "processed_restaurants": [],
                "completed_pages": []
            }
        }
        if not os.path.exists(self.CURRENT_PROGRESS_FILE):
            print(f"No current progress file found, initializing {self.CURRENT_PROGRESS_FILE}")
            self.save_current_progress(default_progress)
            return default_progress
        
        try:
            with open(self.CURRENT_PROGRESS_FILE, 'r', encoding='utf-8') as f:
                progress = json.load(f)
            if not isinstance(progress, dict) or "current_progress" not in progress:
                print(f"Invalid current progress file, resetting to default")
                logging.warning(f"Invalid current progress file structure")
                self.save_current_progress(default_progress)
                return default_progress
            progress["current_progress"]["completed_pages"] = sorted(list(set(
                int(page) for page in progress["current_progress"].get("completed_pages", [])
                if isinstance(page, (int, float)) and page >= 1
            )))
            print(f"Loaded current progress from {self.CURRENT_PROGRESS_FILE}")
            logging.info(f"Loaded current progress: {json.dumps(progress, ensure_ascii=False)}")
            return progress
        except Exception as e:
            print(f"Error loading current progress: {e}")
            logging.error(f"Error loading current progress: {e}")
            self.save_current_progress(default_progress)
            return default_progress

    def save_current_progress(self, progress: Dict = None):
        if progress is None:
            progress = self.current_progress
        temp_filename = None
        try:
            progress["last_updated"] = datetime.now().isoformat()
            if "current_progress" in progress:
                progress["current_progress"]["completed_pages"] = sorted(list(set(
                    int(page) for page in progress["current_progress"].get("completed_pages", [])
                    if isinstance(page, (int, float)) and page >= 1
                )))
            json.dumps(progress, ensure_ascii=False)
            with tempfile.NamedTemporaryFile('w', delete=False, dir='.') as temp_file:
                json.dump(progress, temp_file, indent=2, ensure_ascii=False)
                temp_file.flush()
                os.fsync(temp_file.fileno())
                temp_filename = temp_file.name
            os.replace(temp_filename, self.CURRENT_PROGRESS_FILE)
            print(f"Saved current progress to {self.CURRENT_PROGRESS_FILE}")
            logging.info(f"Saved current progress: {json.dumps(progress, ensure_ascii=False)}")
        except Exception as e:
            print(f"Failed to save current progress: {e}")
            logging.error(f"Failed to save current progress: {e}")
        finally:
            if temp_filename and os.path.exists(temp_filename):
                try:
                    os.remove(temp_filename)
                    print(f"Cleaned up temporary file: {temp_filename}")
                except Exception as e:
                    print(f"Failed to clean up temporary file {temp_filename}: {e}")

    def load_scraped_progress(self) -> Dict:
        default_progress = {
            "completed_areas": [],
            "current_area_index": 0,
            "last_updated": None,
            "all_results": {},
            "current_progress": {
                "area_name": None,
                "current_page": 0,
                "total_pages": 0,
                "current_restaurant": 0,
                "total_restaurants": 0,
                "processed_restaurants": [],
                "completed_pages": []
            }
        }
        if not os.path.exists(self.SCRAPED_PROGRESS_FILE):
            print(f"No scraped progress file found, initializing {self.SCRAPED_PROGRESS_FILE}")
            self.save_scraped_progress(default_progress)
            return default_progress
        
        try:
            with open(self.SCRAPED_PROGRESS_FILE, 'r', encoding='utf-8') as f:
                progress = json.load(f)
            if not isinstance(progress, dict) or "current_progress" not in progress or "all_results" not in progress:
                print(f"Invalid scraped progress file, resetting to default")
                logging.warning(f"Invalid scraped progress file structure")
                self.save_scraped_progress(default_progress)
                return default_progress
            progress["current_progress"]["processed_restaurants"] = list(set(
                str(item) for item in progress["current_progress"].get("processed_restaurants", [])
            ))
            progress["current_progress"]["completed_pages"] = sorted(list(set(
                int(page) for page in progress["current_progress"].get("completed_pages", [])
                if isinstance(page, (int, float)) and page >= 1
            )))
            print(f"Loaded scraped progress from {self.SCRAPED_PROGRESS_FILE}")
            logging.info(f"Loaded scraped progress: {json.dumps(progress, ensure_ascii=False)}")
            return progress
        except Exception as e:
            print(f"Error loading scraped progress: {e}")
            logging.error(f"Error loading scraped progress: {e}")
            self.save_scraped_progress(default_progress)
            return default_progress

    def save_scraped_progress(self, progress: Dict = None):
        if progress is None:
            progress = self.scraped_progress
        temp_filename = None
        try:
            progress["last_updated"] = datetime.now().isoformat()
            if "current_progress" in progress:
                progress["current_progress"]["completed_pages"] = sorted(list(set(
                    int(page) for page in progress["current_progress"].get("completed_pages", [])
                    if isinstance(page, (int, float)) and page >= 1
                )))
            content_str = json.dumps(progress, ensure_ascii=False)
            logging.debug(f"Saving scraped_progress content: {content_str}")
            json.dumps(progress, ensure_ascii=False)
            with tempfile.NamedTemporaryFile('w', delete=False, dir='.') as temp_file:
                json.dump(progress, temp_file, indent=2, ensure_ascii=False)
                temp_file.flush()
                os.fsync(temp_file.fileno())
                temp_filename = temp_file.name
            os.replace(temp_filename, self.SCRAPED_PROGRESS_FILE)
            with open(self.SCRAPED_PROGRESS_FILE, 'r', encoding='utf-8') as f:
                written_content = f.read()
            logging.debug(f"Verified scraped_progress.json content after write: {written_content}")
            mtime = os.path.getmtime(self.SCRAPED_PROGRESS_FILE)
            print(f"Saved scraped progress to {self.SCRAPED_PROGRESS_FILE} at {datetime.fromtimestamp(mtime).isoformat()}")
            logging.info(f"Saved scraped progress to {self.SCRAPED_PROGRESS_FILE}")
        except Exception as e:
            print(f"Failed to save scraped progress: {e}")
            logging.error(f"Failed to save scraped progress: {e}")
        finally:
            if temp_filename and os.path.exists(temp_filename):
                try:
                    os.remove(temp_filename)
                    print(f"Cleaned up temporary file: {temp_filename}")
                except Exception as e:
                    print(f"Failed to clean up temporary file {temp_filename}: {e}")

    def clear_log_file(self):
        try:
            with open('scraper.log', 'w'):
                pass
            logging.info("Cleared scraper.log")
            print("Cleared scraper.log")
        except Exception as e:
            print(f"Failed to clear log file: {e}")
            logging.error(f"Failed to clear log file: {e}")

    def print_progress_details(self):
        try:
            with open(self.CURRENT_PROGRESS_FILE, 'r', encoding='utf-8') as f:
                current = json.load(f)
            print("\nCurrent Progress:")
            print(json.dumps(current, indent=2, ensure_ascii=False))
        except Exception as e:
            print(f"Error printing current progress: {e}")
            logging.error(f"Error printing current progress: {e}")

    async def scrape_and_save_area(self, area_name: str, area_url: str) -> List[Dict]:
        print(f"\n{'='*50}")
        print(f"SCRAPING AREA: {area_name}")
        print(f"URL: {area_url}")
        print(f"{'='*50}\n")
        
        # Checkpoint at the start
        self.save_current_progress()
        self.save_scraped_progress()
        self.commit_progress(f"Started scraping area {area_name}")
        
        all_area_results = self.scraped_progress["all_results"].get(area_name, [])
        current_progress = self.current_progress["current_progress"]
        scraped_current_progress = self.scraped_progress["current_progress"]
        
        is_resuming = current_progress["area_name"] == area_name
        start_page = current_progress["current_page"] if is_resuming else 1
        start_restaurant = current_progress["current_restaurant"] if is_resuming else 0
        
        if is_resuming:
            print(f"Resuming area {area_name} from page {start_page} restaurant {start_restaurant + 1 if start_restaurant > 0 else 1}")
        else:
            current_progress.update({
                "area_name": area_name,
                "current_page": start_page,
                "total_pages": 0,
                "current_restaurant": 0,
                "total_restaurants": 0,
                "processed_restaurants": [],
                "completed_pages": []
            })
            scraped_current_progress.update(current_progress)
            self.save_current_progress()
            self.save_scraped_progress()
            self.commit_progress(f"Started scraping area {area_name}")
        
        skip_categories = {"Grocery, Convenience Store", "Pharmacy", "Flowers", "Electronics", "Grocery, Hypermarket"}
        
        if current_progress["total_pages"] == 0:
            total_pages = await self.determine_total_pages(area_url)
            current_progress["total_pages"] = total_pages
            scraped_current_progress["total_pages"] = total_pages
            self.save_current_progress()
            self.save_scraped_progress()
            self.commit_progress(f"Determined {total_pages} pages for {area_name}")
        else:
            total_pages = current_progress["total_pages"]
        
        print(f"Total pages for {area_name}: {total_pages}")
        
        detailed_csv_filename = os.path.join(self.output_dir, f"{area_name}_detailed.csv")
        
        for page_num in range(start_page, total_pages + 1):
            # Checkpoint before processing page
            self.save_current_progress()
            self.save_scraped_progress()
            self.commit_progress(f"Starting page {page_num} in {area_name}")
            
            if page_num in current_progress["completed_pages"]:
                print(f"Skipping completed page {page_num}")
                continue
            
            page_url = area_url if page_num == 1 else (
                re.sub(r'page=\d+', f'page={page_num}', area_url) if "page=" in area_url else
                f"{area_url}{'&' if '?' in area_url else '?'}page={page_num}"
            )
            
            print(f"\n--- Processing Page {page_num}/{total_pages} for {area_name} ---")
            current_progress["current_page"] = page_num
            scraped_current_progress["current_page"] = page_num
            self.save_current_progress()
            self.save_scraped_progress()
            
            max_retries = 3
            for attempt in range(max_retries):
                try:
                    restaurants_on_page = await self.get_page_restaurants(page_url, page_num)
                    if not restaurants_on_page:
                        raise Exception("No restaurants found")
                    print(f"Found {len(restaurants_on_page)} restaurants on page {page_num}")
                    break
                except Exception as e:
                    print(f"Error on page {page_num}: {e}")
                    logging.error(f"Error on page {page_num}: {e}")
                    if attempt < max_retries - 1:
                        print(f"Retrying ({attempt + 1}/{max_retries})...")
                        await asyncio.sleep(5)
                    else:
                        print(f"Skipping page {page_num} after {max_retries} attempts")
                        restaurants_on_page = []
            
            if current_progress["total_restaurants"] == 0 or page_num > start_page:
                current_progress["total_restaurants"] = len(restaurants_on_page)
                scraped_current_progress["total_restaurants"] = len(restaurants_on_page)
                if not is_resuming or page_num > start_page:
                    current_progress["current_restaurant"] = 0
                    scraped_current_progress["current_restaurant"] = 0
            
            page_restaurants = []
            for rest_idx, restaurant in enumerate(restaurants_on_page):
                rest_num = rest_idx + 1
                restaurant_name = restaurant.get("name", "").strip()
                
                if rest_num <= current_progress["current_restaurant"]:
                    print(f"Skipping processed restaurant {rest_num}/{len(restaurants_on_page)}: {restaurant_name}")
                    continue
                
                is_already_processed = any(
                    r.get("name", "").strip() == restaurant_name and r.get("page", 0) == page_num
                    for r in all_area_results
                ) or restaurant_name in current_progress["processed_restaurants"]
                
                if is_already_processed:
                    print(f"Skipping restaurant {rest_num}/{len(restaurants_on_page)}: {restaurant_name} - Already processed")
                    current_progress["current_restaurant"] = rest_num
                    scraped_current_progress["current_restaurant"] = rest_num
                    self.save_current_progress()
                    self.save_scraped_progress()
                    continue
                
                current_progress["current_restaurant"] = rest_num
                scraped_current_progress["current_restaurant"] = rest_num
                
                if any(category in restaurant['cuisine'] for category in skip_categories):
                    print(f"\nSkipping restaurant {rest_num}/{len(restaurants_on_page)}: {restaurant_name} - Category: {restaurant['cuisine']}")
                    if restaurant_name and restaurant_name not in current_progress["processed_restaurants"]:
                        current_progress["processed_restaurants"].append(restaurant_name)
                        scraped_current_progress["processed_restaurants"].append(restaurant_name)
                    self.save_current_progress()
                    self.save_scraped_progress()
                    continue
                
                print(f"\nProcessing restaurant {rest_num}/{len(restaurants_on_page)} on page {page_num}: {restaurant_name}")
                
                try:
                    restaurant.setdefault("menu_items", {})
                    restaurant.setdefault("info", {})
                    restaurant.setdefault("reviews", {})
                    restaurant["page"] = page_num
                    
                    async def timeout_task(task, timeout=60):
                        try:
                            return await asyncio.wait_for(task, timeout=timeout)
                        except asyncio.TimeoutError:
                            print(f"Timeout while processing task for {restaurant_name}")
                            logging.error(f"Timeout while processing task for {restaurant_name}")
                            return None
                    
                    print(f"Fetching menu for {restaurant_name}...")
                    menu_data = await timeout_task(self.talabat_scraper.get_restaurant_menu(restaurant['url']))
                    if menu_data:
                        restaurant['menu_items'] = menu_data
                    else:
                        print(f"No menu data retrieved for {restaurant_name}")
                    
                    print(f"Fetching info for {restaurant_name}...")
                    info_data = await timeout_task(self.talabat_scraper.get_restaurant_info(restaurant['url']))
                    if info_data:
                        restaurant['info'] = info_data
                    else:
                        print(f"No info data retrieved for {restaurant_name}")
                    
                    if restaurant['info'].get('Reviews URL') and restaurant['info']['Reviews URL'] != 'Not Available':
                        print(f"Fetching reviews for {restaurant_name}...")
                        reviews_data = self.talabat_scraper.get_reviews_data(restaurant['info']['Reviews URL'])
                        restaurant['reviews'] = reviews_data or {}
                    else:
                        print(f"No reviews URL available for {restaurant_name}")
                    
                    page_restaurants.append(restaurant)
                    all_area_results.append(restaurant)
                    if restaurant_name and restaurant_name not in current_progress["processed_restaurants"]:
                        current_progress["processed_restaurants"].append(restaurant_name)
                        scraped_current_progress["processed_restaurants"].append(restaurant_name)
                    self.scraped_progress["all_results"][area_name] = all_area_results
                    logging.debug(f"Updated all_results for {area_name}: {len(all_area_results)} restaurants")
                    
                    self.save_current_progress()
                    self.save_scraped_progress()
                    
                    await asyncio.sleep(2)
                
                except Exception as e:
                    print(f"Error processing restaurant {rest_num}/{len(restaurants_on_page)}: {restaurant_name}: {e}")
                    logging.error(f"Error processing restaurant {restaurant_name}: {e}")
                    import traceback
                    traceback.print_exc()
                    if restaurant_name and restaurant_name not in current_progress["processed_restaurants"]:
                        current_progress["processed_restaurants"].append(restaurant_name)
                        scraped_current_progress["processed_restaurants"].append(restaurant_name)
                    self.save_current_progress()
                    self.save_scraped_progress()
            
            # Save JSON after processing all restaurants on the page
            try:
                json_filename = os.path.join(self.output_dir, f"{area_name}.json")
                with open(json_filename, 'w', encoding='utf-8') as f:
                    json.dump(all_area_results, f, indent=2, ensure_ascii=False)
                logging.info(f"Saved {len(all_area_results)} restaurants to {json_filename}")
            except Exception as e:
                print(f"Failed to save JSON for {area_name}: {e}")
                logging.error(f"Failed to save JSON for {area_name}: {e}")
            
            # Save restaurants for the page to detailed CSV
            if page_restaurants:
                try:
                    print(f"Saving {len(page_restaurants)} restaurants from page {page_num} to {detailed_csv_filename}")
                    self.create_detailed_excel_sheet(area_name, page_restaurants, detailed_csv_filename)
                except Exception as e:
                    print(f"Failed to save detailed CSV for page {page_num}: {e}")
                    logging.error(f"Failed to save detailed CSV for page {page_num}: {e}")
            
            # Clear log file
            self.clear_log_file()
            
            # Mark page as complete
            if page_num not in current_progress["completed_pages"]:
                current_progress["completed_pages"].append(page_num)
                scraped_current_progress["completed_pages"].append(page_num)
            current_progress["current_restaurant"] = 0
            scraped_current_progress["current_restaurant"] = 0
            self.save_current_progress()
            self.save_scraped_progress()
            self.commit_progress(f"Completed page {page_num} in {area_name}")
            await asyncio.sleep(3)
        
        # Final JSON save
        json_filename = os.path.join(self.output_dir, f"{area_name}.json")
        with open(json_filename, 'w', encoding='utf-8') as f:
            json.dump(all_area_results, f, indent=2, ensure_ascii=False)
        logging.info(f"Final save: {len(all_area_results)} restaurants to {json_filename}")
        
        # Create simplified Excel workbook
        simplified_workbook = Workbook()
        self.create_excel_sheet(simplified_workbook, area_name, all_area_results)
        simplified_excel_filename = os.path.join(self.output_dir, f"{area_name}.xlsx")
        simplified_workbook.save(simplified_excel_filename)
        print(f"Simplified Excel file saved: {simplified_excel_filename}")
        
        # Upload both files to Google Drive
        if self.upload_to_drive(simplified_excel_filename):
            print(f"Uploaded {simplified_excel_filename} to Google Drive")
        else:
            print(f"Failed to upload {simplified_excel_filename} to Google Drive")
        
        if self.upload_to_drive(detailed_csv_filename):
            print(f"Uploaded {detailed_csv_filename} to Google Drive")
        else:
            print(f"Failed to upload {detailed_csv_filename} to Google Drive")
        
        current_progress.update({
            "area_name": None,
            "current_page": 0,
            "total_pages": 0,
            "current_restaurant": 0,
            "total_restaurants": 0,
            "processed_restaurants": [],
            "completed_pages": []
        })
        scraped_current_progress.update(current_progress)
        self.save_current_progress()
        self.save_scraped_progress()
        self.print_progress_details()
        self.commit_progress(f"Completed area {area_name}")
        
        print(f"Saved {len(all_area_results)} restaurants for {area_name}")
        return all_area_results

    def commit_progress(self, message: str):
        try:
            status_result = subprocess.run(["git", "status", "--porcelain"], capture_output=True, text=True, check=True)
            logging.debug(f"Git status before staging: {status_result.stdout}")
            
            subprocess.run(["git", "add", self.CURRENT_PROGRESS_FILE], check=True)
            subprocess.run(["git", "add", self.SCRAPED_PROGRESS_FILE], check=True)
            subprocess.run(["git", "add", self.output_dir], check=True)
            
            diff_result = subprocess.run(["git", "diff", "--staged"], capture_output=True, text=True, check=True)
            logging.debug(f"Git diff --staged: {diff_result.stdout}")
            
            result = subprocess.run(["git", "commit", "-m", message], capture_output=True, text=True)
            if result.returncode == 0:
                print(f"Committed progress: {message}")
                logging.info(f"Committed progress: {message}")
            else:
                print(f"No changes to commit for: {message}")
                logging.warning(f"No changes to commit: {result.stderr}")
            
            push_result = subprocess.run(["git", "push"], capture_output=True, text=True)
            if push_result.returncode == 0:
                print(f"Pushed progress: {message}")
                logging.info(f"Pushed progress: {message}")
            else:
                print(f"Failed to push progress: {push_result.stderr}")
                logging.error(f"Failed to push progress: {push_result.stderr}")
            
            # Clean git temporary files
            gc_result = subprocess.run(["git", "gc", "--prune=now"], capture_output=True, text=True)
            if gc_result.returncode == 0:
                print("Cleaned git temporary files")
                logging.info("Cleaned git temporary files")
            else:
                print(f"Failed to clean git temporary files: {gc_result.stderr}")
                logging.error(f"Failed to clean git temporary files: {gc_result.stderr}")
        
        except subprocess.CalledProcessError as e:
            print(f"Failed to commit progress: {e}")
            logging.error(f"Failed to commit progress: {e}")

    async def determine_total_pages(self, area_url: str) -> int:
        print(f"Determining total pages for URL: {area_url}")
        try:
            async with async_playwright() as p:
                browser = await p.firefox.launch(headless=True)
                context = await browser.new_context(
                    viewport={'width': 1920, 'height': 1080},
                    user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                )
                page = await context.new_page()
                page.set_default_timeout(120000)
                
                response = await page.goto(area_url, wait_until='domcontentloaded')
                if not response or not response.ok:
                    print(f"Failed to load page: {response.status if response else 'No response'}")
                    return 1
                
                await page.wait_for_selector("ul[data-test='pagination'], .vendor-card, [data-testid='restaurant-a']", timeout=30000)
                
                last_page = 1
                pagination = await page.query_selector("ul[data-test='pagination']")
                if pagination:
                    items = await pagination.query_selector_all("li[data-testid='paginate-link']")
                    if items and len(items) > 1:
                        last_page_item = items[-2]
                        last_page_link = await last_page_item.query_selector("a[page]")
                        if last_page_link:
                            last_page_attr = await last_page_link.get_attribute("page")
                            if last_page_attr and last_page_attr.isdigit():
                                last_page = int(last_page_attr)
                
                await browser.close()
                return last_page
        except Exception as e:
            print(f"Error determining total pages: {e}")
            return 1

    async def get_page_restaurants(self, page_url: str, page_num: int) -> List[Dict]:
        browser = None
        try:
            async with async_playwright() as p:
                browser = await p.firefox.launch(headless=True)
                context = await browser.new_context(
                    viewport={'width': 1920, 'height': 1080},
                    user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                )
                page = await context.new_page()
                page.set_default_timeout(120000)
                
                response = await page.goto(page_url, wait_until='domcontentloaded')
                if not response or not response.ok:
                    print(f"Failed to load page {page_num}: {response.status if response else 'No response'}")
                    return []
                
                await page.wait_for_selector(".vendor-card, [data-testid='restaurant-a']", timeout=30000)
                return await self.talabat_scraper._extract_restaurants_from_page(page, page_num)
        except Exception as e:
            print(f"Error getting page restaurants: {e}")
            import traceback
            traceback.print_exc()
            return []
        finally:
            if browser:
                await browser.close()

    def create_excel_sheet(self, workbook, sheet_name: str, data: List[Dict]):
        sheet = workbook.create_sheet(title=sheet_name)
        try:
            simplified_data = []
            for restaurant in data:
                restaurant_info = {
                    "Name": restaurant.get("name", ""),
                    "Cuisine": restaurant.get("cuisine", ""),
                    "Rating": restaurant.get("rating", ""),
                    "Delivery Time": restaurant.get("delivery_time", ""),
                    "Delivery Fee": restaurant.get("delivery_fee", ""),
                    "Min Order": restaurant.get("min_order", ""),
                    "URL": restaurant.get("url", ""),
                }
                if restaurant.get("info"):
                    restaurant_info.update({
                        "Address": restaurant["info"].get("Address", ""),
                        "Working Hours": restaurant["info"].get("Working Hours", ""),
                    })
                if restaurant.get("reviews") and restaurant["reviews"].get("Rating_value"):
                    restaurant_info.update({
                        "Rating Value": restaurant["reviews"]["Rating_value"],
                        "Ratings Count": restaurant["reviews"].get("Ratings_count", ""),
                        "Reviews Count": restaurant["reviews"].get("Reviews_count", ""),
                    })
                if restaurant.get("menu_items"):
                    restaurant_info["Menu Categories"] = len(restaurant["menu_items"])
                    item_count = sum(len(items) for items in restaurant["menu_items"].values())
                    restaurant_info["Menu Items"] = item_count
                simplified_data.append(restaurant_info)
            
            if simplified_data:
                df = pd.DataFrame(simplified_data)
                for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        sheet.cell(row=r_idx, column=c_idx, value=value)
                for column in sheet.columns:
                    max_length = max(len(str(cell.value or "")) for cell in column)
                    column_letter = get_column_letter(column[0].column)
                    sheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
            else:
                sheet.cell(row=1, column=1, value="No data found for this area")
        except Exception as e:
            print(f"Error creating Excel sheet for {sheet_name}: {str(e)}")
            sheet.cell(row=1, column=1, value=f"Error processing data: {str(e)}")

    def flatten_menu_items(self, menu_items):
        if not isinstance(menu_items, dict):
            return ''
        items = []
        for category, dishes in menu_items.items():
            for dish in dishes:
                if isinstance(dish, dict):
                    name = dish.get('name', '')
                    price = dish.get('prices', {}).get('new_price', '')
                    items.append(f"{name}: {price}")
        return '; '.join(items)

    def create_detailed_excel_sheet(self, area_name: str, data: List[Dict], excel_filename: str):
        columns = [
            'restaurant name', 'cuisine', 'restaurant url', 'general rating',
            'restaurant in page number', 'delivery time', 'delivery fees',
            'minimum order', 'tracking status', 'contactless', 'menu items',
            'Address', 'Reviews URL', 'Pre-Order status', 'Payment types',
            'reviews rating value', 'reviews Ratings count', 'reviews count',
            'General review', 'Order Packaging reviews', 'Value for money reviews',
            'Delivery time reviews', 'Quality of food reviews', 'Customer reviews'
        ]
        
        try:
            rows = []
            for restaurant in data:
                row = {
                    'restaurant name': restaurant.get('name', ''),
                    'cuisine': restaurant.get('cuisine', ''),
                    'restaurant url': restaurant.get('url', ''),
                    'general rating': restaurant.get('rating', ''),
                    'restaurant in page number': restaurant.get('page', ''),
                    'delivery time': restaurant.get('delivery_time', ''),
                    'delivery fees': restaurant.get('delivery_fee', ''),
                    'minimum order': restaurant.get('min_order', ''),
                    'tracking status': restaurant.get('tracking_status', ''),
                    'contactless': restaurant.get('contactless', ''),
                    'menu items': self.flatten_menu_items(restaurant.get('menu_items', {})),
                    'Address': restaurant.get('info', {}).get('Address', ''),
                    'Reviews URL': restaurant.get('info', {}).get('Reviews URL', ''),
                    'Pre-Order status': restaurant.get('info', {}).get('Pre-Order', ''),
                    'Payment types': ', '.join(restaurant.get('info', {}).get('Payment', [])),
                    'reviews rating value': restaurant.get('reviews', {}).get('Rating_value', ''),
                    'reviews Ratings count': restaurant.get('reviews', {}).get('Ratings_count', ''),
                    'reviews count': restaurant.get('reviews', {}).get('Reviews_count', ''),
                    'General review': '; '.join(restaurant.get('reviews', {}).get('General_review', [])),
                    'Order Packaging reviews': restaurant.get('reviews', {}).get('Specific_reviews', {}).get('Order Packaging', ''),
                    'Value for money reviews': restaurant.get('reviews', {}).get('Specific_reviews', {}).get('Value for money', ''),
                    'Delivery time reviews': restaurant.get('reviews', {}).get('Specific_reviews', {}).get('Delivery time', ''),
                    'Quality of food reviews': restaurant.get('reviews', {}).get('Specific_reviews', {}).get('Quality of food', ''),
                    'Customer reviews': '; '.join([
                        f"{rev.get('reviewer_name', 'Unknown')}: {rev.get('review_comment', '')} ({rev.get('review_date', '')})"
                        for rev in restaurant.get('reviews', {}).get('Customer_reviews', [])
                        if isinstance(rev, dict)
                    ])
                }
                rows.append(row)
            
            if not rows:
                print(f"No data to save for {area_name}")
                return
            
            csv_filename = excel_filename
            new_df = pd.DataFrame(rows, columns=columns)
            new_df.to_csv(csv_filename, index=False, encoding='utf-8')
            print(f"Saved {len(rows)} restaurants to {csv_filename}")
        
        except Exception as e:
            print(f"Error saving detailed CSV for {area_name}: {str(e)}")
            logging.error(f"Error saving detailed CSV for {area_name}: {str(e)}")

    @retry(tries=3, delay=2, backoff=2)
    def upload_to_drive(self, file_path):
        print(f"\nUploading {file_path} to Google Drive...")
        try:
            credentials_json = os.environ.get('TALABAT_GCLOUD_KEY_JSON')
            if not credentials_json:
                print("Error: TALABAT_GCLOUD_KEY_JSON environment variable is empty or not set!")
                return False
            self.drive_uploader = SavingOnDrive(credentials_json=credentials_json)
            if not self.drive_uploader.authenticate():
                print("Failed to authenticate with Google Drive. Check TALABAT_GCLOUD_KEY_JSON validity.")
                return False
            file_ids = self.drive_uploader.upload_to_multiple_folders(file_path)
            success = len(file_ids) == 2
            if success:
                print(f"Successfully uploaded {file_path} to Google Drive")
            else:
                print(f"Failed to upload {file_path}: Incomplete upload to folders")
            return success
        except Exception as e:
            print(f"Error uploading to Google Drive: {str(e)}")
            return False

    async def run(self):
        ahmadi_areas = [
            ("الظهر", "https://www.talabat.com/kuwait/restaurants/59/dhaher"),
            ("الرقه", "https://www.talabat.com/kuwait/restaurants/37/riqqa"),
            ("هدية", "https://www.talabat.com/kuwait/restaurants/30/hadiya"),
            ("المنقف", "https://www.talabat.com/kuwait/restaurants/32/mangaf"),
            ("أبو حليفة", "https://www.talabat.com/kuwait/restaurants/2/abu-halifa"),
            ("الفنطاس", "https://www.talabat.com/kuwait/restaurants/38/fintas"),
            ("العقيلة", "https://www.talabat.com/kuwait/restaurants/79/egaila"),
            ("الصباحية", "https://www.talabat.com/kuwait/restaurants/31/sabahiya"),
            ("الأحمدي", "https://www.talabat.com/kuwait/restaurants/3/al-ahmadi"),
            ("الفحيحيل", "https://www.talabat.com/kuwait/restaurants/5/fahaheel"),
            ("شرق الأحمدي", "https://www.talabat.com/kuwait/restaurants/3/al-ahmadi"),
            ("ضاحية علي صباح السالم", "https://www.talabat.com/kuwait/restaurants/82/ali-sabah-al-salem-umm-al-hayman"),
            ("ميناء عبد الله", "https://www.talabat.com/kuwait/restaurants/100/mina-abdullah"),
            ("بنيدر", "https://www.talabat.com/kuwait/restaurants/6650/bnaider"),
            ("الزور", "https://www.talabat.com/kuwait/restaurants/2053/zour"),
            ("الجليعة", "https://www.talabat.com/kuwait/restaurants/6860/al-julaiaa"),
            ("المهبولة", "https://www.talabat.com/kuwait/restaurants/24/mahboula"),
            ("النويصيب", "https://www.talabat.com/kuwait/restaurants/2054/nuwaiseeb"),
            ("الخيران", "https://www.talabat.com/kuwait/restaurants/2726/khairan"),
            ("الوفرة", "https://www.talabat.com/kuwait/restaurants/2057/wafra-farms"),
            ("ضاحية فهد الأحمد", "https://www.talabat.com/kuwait/restaurants/98/fahad-al-ahmed"),
            ("ضاحية جابر العلي", "https://www.talabat.com/kuwait/restaurants/60/jaber-al-ali"),
            ("مدينة صباح الأحمد السكنية", "https://www.talabat.com/kuwait/restaurants/6931/sabah-al-ahmad-2"),
            ("مدينة صباح الأحمد البحرية", "https://www.talabat.com/kuwait/restaurants/2726/khairan"),
            ("ميناء الأحمدي", "https://www.talabat.com/kuwait/restaurants/3/al-ahmadi")
        ]
        
        simplified_excel_filename = os.path.join(self.output_dir, "الاحمدي.xlsx")
        simplified_workbook = Workbook()
        if "Sheet" in simplified_workbook.sheetnames:
            simplified_workbook.remove(simplified_workbook["Sheet"])
        
        completed_areas = self.current_progress["completed_areas"]
        current_area_index = self.current_progress["current_area_index"]
        
        print(f"Starting from area index {current_area_index}")
        print(f"Already completed areas: {', '.join(completed_areas) if completed_areas else 'None'}")
        
        resuming_area = self.current_progress["current_progress"]["area_name"]
        if resuming_area:
            for idx, (area_name, _) in enumerate(ahmadi_areas):
                if area_name == resuming_area:
                    print(f"Resuming from area {resuming_area} (index {idx})")
                    current_area_index = idx
                    self.current_progress["current_area_index"] = idx
                    self.scraped_progress["current_area_index"] = idx
                    self.save_current_progress()
                    self.save_scraped_progress()
                    self.commit_progress(f"Resuming from area {resuming_area}")
                    break
        
        for idx, (area_name, area_url) in enumerate(ahmadi_areas):
            if area_name in completed_areas and area_name != resuming_area:
                print(f"Skipping completed area: {area_name}")
                continue
            if idx < current_area_index:
                print(f"Skipping area {area_name} (index {idx} < {current_area_index})")
                continue
            
            self.current_progress["current_area_index"] = idx
            self.scraped_progress["current_area_index"] = idx
            self.save_current_progress()
            self.save_scraped_progress()
            self.commit_progress(f"Starting area {area_name} at index {idx}")
            
            try:
                area_results = await self.scrape_and_save_area(area_name, area_url)
                self.create_excel_sheet(simplified_workbook, area_name, area_results)
                simplified_workbook.save(simplified_excel_filename)
                print(f"Updated simplified Excel file: {simplified_excel_filename}")
                
                if area_name not in completed_areas:
                    completed_areas.append(area_name)
                    self.current_progress["completed_areas"] = completed_areas
                    self.scraped_progress["completed_areas"] = completed_areas
                self.save_current_progress()
                self.save_scraped_progress()
                self.print_progress_details()
                self.commit_progress(f"Completed area {area_name} in run")
                await asyncio.sleep(5)
            
            except Exception as e:
                print(f"Error processing area {area_name}: {e}")
                logging.error(f"Error processing area {area_name}: {e}")
                import traceback
                traceback.print_exc()
                self.save_current_progress()
                self.save_scraped_progress()
                self.commit_progress(f"Progress update after error in {area_name}")
        
        simplified_workbook.save(simplified_excel_filename)
        combined_json_filename = os.path.join(self.output_dir, "الاحمدي_all.json")
        with open(combined_json_filename, 'w', encoding='utf-8') as f:
            json.dump(self.scraped_progress["all_results"], f, indent=2, ensure_ascii=False)
        
        print(f"\n{'='*50}")
        print(f"SCRAPING COMPLETED")
        print(f"Simplified Excel file saved: {simplified_excel_filename}")
        print(f"Combined JSON saved: {combined_json_filename}")
        
        if len(completed_areas) == len(ahmadi_areas):
            if self.upload_to_drive(simplified_excel_filename):
                print(f"Uploaded simplified Excel file to Google Drive")
            else:
                print(f"Failed to upload simplified Excel file to Google Drive")
        else:
            print(f"Scraping incomplete ({len(completed_areas)}/{len(ahmadi_areas)} areas)")
        
        self.commit_progress("Final progress update after run")

async def main():
    try:
        scraper = MainScraper()
        await scraper.run()
    except KeyboardInterrupt:
        print("\nInterrupted. Saving progress...")
        if 'scraper' in locals():
            scraper.save_current_progress()
            scraper.save_scraped_progress()
            scraper.commit_progress("Progress saved after interruption")
        print("Progress saved. Exiting.")
    except Exception as e:
        print(f"Critical error: {e}")
        import traceback
        traceback.print_exc()
        if 'scraper' in locals():
            scraper.save_current_progress()
            scraper.save_scraped_progress()
            scraper.commit_progress("Progress saved after critical error")
        sys.exit(1)

if __name__ == "__main__":
    scraper = MainScraper()
    scraper.print_progress_details()
    asyncio.run(scraper.run())




# import asyncio
# import json
# import os
# import tempfile
# import sys
# import subprocess
# from retry import retry
# import re
# from typing import Dict, List, Tuple
# import pandas as pd
# from openpyxl import Workbook, load_workbook
# from openpyxl.utils.dataframe import dataframe_to_rows
# from openpyxl.utils import get_column_letter
# from playwright.async_api import async_playwright
# from talabat_main_scraper import TalabatScraper
# from SavingOnDrive import SavingOnDrive
# from time import sleep
# from datetime import datetime
# import logging

# logging.basicConfig(
#     filename='scraper.log',
#     level=logging.DEBUG,
#     format='%(asctime)s - %(levelname)s - %(message)s'
# )

# class MainScraper:
#     CURRENT_PROGRESS_FILE = "current_progress.json"
#     SCRAPED_PROGRESS_FILE = "scraped_progress.json"

#     def __init__(self):
#         self.talabat_scraper = TalabatScraper()
#         self.output_dir = "output"
#         credentials_json = os.environ.get('TALABAT_GCLOUD_KEY_JSON')
#         self.drive_uploader = SavingOnDrive(credentials_json=credentials_json)
        
#         os.makedirs(self.output_dir, exist_ok=True)
        
#         self.current_progress = self.load_current_progress()
#         self.scraped_progress = self.load_scraped_progress()
        
#         self.github_token = os.environ.get('GITHUB_TOKEN')
#         self.ensure_playwright_browsers()

#     def ensure_playwright_browsers(self):
#         try:
#             print("Installing Playwright browsers...")
#             subprocess.run([sys.executable, "-m", "playwright", "install", "chromium", "firefox"], 
#                           check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
#             print("Playwright browsers installed successfully")
#         except subprocess.CalledProcessError as e:
#             print(f"Error installing Playwright browsers: {e}")
#             logging.error(f"Error installing Playwright browsers: {e}")

#     def load_current_progress(self) -> Dict:
#         default_progress = {
#             "completed_areas": [],
#             "current_area_index": 0,
#             "last_updated": None,
#             "current_progress": {
#                 "area_name": None,
#                 "current_page": 0,
#                 "total_pages": 0,
#                 "current_restaurant": 0,
#                 "total_restaurants": 0,
#                 "processed_restaurants": [],
#                 "completed_pages": []
#             }
#         }
#         if not os.path.exists(self.CURRENT_PROGRESS_FILE):
#             print(f"No current progress file found, initializing {self.CURRENT_PROGRESS_FILE}")
#             self.save_current_progress(default_progress)
#             return default_progress
        
#         try:
#             with open(self.CURRENT_PROGRESS_FILE, 'r', encoding='utf-8') as f:
#                 progress = json.load(f)
#             # Validate structure
#             if not isinstance(progress, dict) or "current_progress" not in progress:
#                 print(f"Invalid current progress file, resetting to default")
#                 logging.warning(f"Invalid current progress file structure")
#                 self.save_current_progress(default_progress)
#                 return default_progress
#             # Keep processed_restaurants as-is, deduplication handled at append
#             progress["current_progress"]["completed_pages"] = sorted(list(set(
#                 int(page) for page in progress["current_progress"].get("completed_pages", [])
#                 if isinstance(page, (int, float)) and page >= 1
#             )))
#             print(f"Loaded current progress from {self.CURRENT_PROGRESS_FILE}")
#             logging.info(f"Loaded current progress: {json.dumps(progress, ensure_ascii=False)}")
#             return progress
#         except Exception as e:
#             print(f"Error loading current progress: {e}")
#             logging.error(f"Error loading current progress: {e}")
#             self.save_current_progress(default_progress)
#             return default_progress

#     def save_current_progress(self, progress: Dict = None):
#         if progress is None:
#             progress = self.current_progress
#         try:
#             progress["last_updated"] = datetime.now().isoformat()
#             # Only deduplicate completed_pages
#             if "current_progress" in progress:
#                 progress["current_progress"]["completed_pages"] = sorted(list(set(
#                     int(page) for page in progress["current_progress"].get("completed_pages", [])
#                     if isinstance(page, (int, float)) and page >= 1
#                 )))
#             # Validate JSON serializability
#             json.dumps(progress, ensure_ascii=False)
#             with tempfile.NamedTemporaryFile('w', delete=False, dir='.') as temp_file:
#                 json.dump(progress, temp_file, indent=2, ensure_ascii=False)
#                 temp_file.flush()
#                 os.fsync(temp_file.fileno())
#                 temp_filename = temp_file.name
#             os.replace(temp_filename, self.CURRENT_PROGRESS_FILE)
#             print(f"Saved current progress to {self.CURRENT_PROGRESS_FILE}")
#             logging.info(f"Saved current progress: {json.dumps(progress, ensure_ascii=False)}")
#         except Exception as e:
#             print(f"Failed to save current progress: {e}")
#             logging.error(f"Failed to save current progress: {e}")
            
#     def load_scraped_progress(self) -> Dict:
#         default_progress = {
#             "completed_areas": [],
#             "current_area_index": 0,
#             "last_updated": None,
#             "all_results": {},
#             "current_progress": {
#                 "area_name": None,
#                 "current_page": 0,
#                 "total_pages": 0,
#                 "current_restaurant": 0,
#                 "total_restaurants": 0,
#                 "processed_restaurants": [],
#                 "completed_pages": []
#             }
#         }
#         if not os.path.exists(self.SCRAPED_PROGRESS_FILE):
#             print(f"No scraped progress file found, initializing {self.SCRAPED_PROGRESS_FILE}")
#             self.save_scraped_progress(default_progress)
#             return default_progress
        
#         try:
#             with open(self.SCRAPED_PROGRESS_FILE, 'r', encoding='utf-8') as f:
#                 progress = json.load(f)
#             # Validate structure
#             if not isinstance(progress, dict) or "current_progress" not in progress or "all_results" not in progress:
#                 print(f"Invalid scraped progress file, resetting to default")
#                 logging.warning(f"Invalid scraped progress file structure")
#                 self.save_scraped_progress(default_progress)
#                 return default_progress
#             # Deduplicate processed_restaurants and completed_pages
#             progress["current_progress"]["processed_restaurants"] = list(set(
#                 str(item) for item in progress["current_progress"].get("processed_restaurants", [])
#             ))
#             progress["current_progress"]["completed_pages"] = sorted(list(set(
#                 int(page) for page in progress["current_progress"].get("completed_pages", [])
#                 if isinstance(page, (int, float)) and page >= 1
#             )))
#             print(f"Loaded scraped progress from {self.SCRAPED_PROGRESS_FILE}")
#             logging.info(f"Loaded scraped progress: {json.dumps(progress, ensure_ascii=False)}")
#             return progress
#         except Exception as e:
#             print(f"Error loading scraped progress: {e}")
#             logging.error(f"Error loading scraped progress: {e}")
#             self.save_scraped_progress(default_progress)
#             return default_progress


#     def print_progress_details(self):
#         try:
#             with open(self.CURRENT_PROGRESS_FILE, 'r', encoding='utf-8') as f:
#                 current = json.load(f)
#             print("\nCurrent Progress:")
#             print(json.dumps(current, indent=2, ensure_ascii=False))
#         except Exception as e:
#             print(f"Error printing current progress: {e}")
#             logging.error(f"Error printing current progress: {e}")


#     async def scrape_and_save_area(self, area_name: str, area_url: str) -> List[Dict]:
#         print(f"\n{'='*50}")
#         print(f"SCRAPING AREA: {area_name}")
#         print(f"URL: {area_url}")
#         print(f"{'='*50}\n")
        
#         all_area_results = self.scraped_progress["all_results"].get(area_name, [])
#         current_progress = self.current_progress["current_progress"]
#         scraped_current_progress = self.scraped_progress["current_progress"]
        
#         is_resuming = current_progress["area_name"] == area_name
#         start_page = current_progress["current_page"] if is_resuming else 1
#         start_restaurant = current_progress["current_restaurant"] if is_resuming else 0
        
#         if is_resuming:
#             print(f"Resuming area {area_name} from page {start_page} restaurant {start_restaurant + 1 if start_restaurant > 0 else 1}")
#         else:
#             current_progress.update({
#                 "area_name": area_name,
#                 "current_page": start_page,
#                 "total_pages": 0,
#                 "current_restaurant": 0,
#                 "total_restaurants": 0,
#                 "processed_restaurants": [],
#                 "completed_pages": []
#             })
#             scraped_current_progress.update(current_progress)
#             self.save_current_progress()
#             self.save_scraped_progress()
#             self.commit_progress(f"Started scraping area {area_name}")
        
#         skip_categories = {"Grocery, Convenience Store", "Pharmacy", "Flowers", "Electronics", "Grocery, Hypermarket"}
        
#         if current_progress["total_pages"] == 0:
#             total_pages = await self.determine_total_pages(area_url)
#             current_progress["total_pages"] = total_pages
#             scraped_current_progress["total_pages"] = total_pages
#             self.save_current_progress()
#             self.save_scraped_progress()
#             self.commit_progress(f"Determined {total_pages} pages for {area_name}")
#         else:
#             total_pages = current_progress["total_pages"]
        
#         print(f"Total pages for {area_name}: {total_pages}")
        
#         detailed_excel_filename = os.path.join(self.output_dir, f"{area_name}_detailed.xlsx")
        
#         for page_num in range(start_page, total_pages + 1):
#             if page_num in current_progress["completed_pages"]:
#                 print(f"Skipping completed page {page_num}")
#                 continue
            
#             page_url = area_url if page_num == 1 else (
#                 re.sub(r'page=\d+', f'page={page_num}', area_url) if "page=" in area_url else
#                 f"{area_url}{'&' if '?' in area_url else '?'}page={page_num}"
#             )
            
#             print(f"\n--- Processing Page {page_num}/{total_pages} for {area_name} ---")
#             current_progress["current_page"] = page_num
#             scraped_current_progress["current_page"] = page_num
#             self.save_current_progress()
#             self.save_scraped_progress()
#             self.commit_progress(f"Started page {page_num} in {area_name}")
            
#             max_retries = 3
#             for attempt in range(max_retries):
#                 try:
#                     restaurants_on_page = await self.get_page_restaurants(page_url, page_num)
#                     if not restaurants_on_page:
#                         raise Exception("No restaurants found")
#                     print(f"Found {len(restaurants_on_page)} restaurants on page {page_num}")
#                     break
#                 except Exception as e:
#                     print(f"Error on page {page_num}: {e}")
#                     logging.error(f"Error on page {page_num}: {e}")
#                     if attempt < max_retries - 1:
#                         print(f"Retrying ({attempt + 1}/{max_retries})...")
#                         await asyncio.sleep(5)
#                     else:
#                         print(f"Skipping page {page_num} after {max_retries} attempts")
#                         restaurants_on_page = []
            
#             if current_progress["total_restaurants"] == 0 or page_num > start_page:
#                 current_progress["total_restaurants"] = len(restaurants_on_page)
#                 scraped_current_progress["total_restaurants"] = len(restaurants_on_page)
#                 if not is_resuming or page_num > start_page:
#                     current_progress["current_restaurant"] = 0
#                     scraped_current_progress["current_restaurant"] = 0
            
#             page_restaurants = []
#             for rest_idx, restaurant in enumerate(restaurants_on_page):
#                 rest_num = rest_idx + 1
#                 restaurant_name = restaurant.get("name", "").strip()
                
#                 # Skip if restaurant is already processed on this page
#                 if rest_num <= current_progress["current_restaurant"]:
#                     print(f"Skipping processed restaurant {rest_num}/{len(restaurants_on_page)}: {restaurant_name}")
#                     continue
                
#                 # Check if restaurant was processed before starting scraping
#                 is_already_processed = any(
#                     r.get("name", "").strip() == restaurant_name and r.get("page", 0) == page_num
#                     for r in all_area_results
#                 ) or restaurant_name in current_progress["processed_restaurants"]
                
#                 if is_already_processed:
#                     print(f"Skipping restaurant {rest_num}/{len(restaurants_on_page)}: {restaurant_name} - Already processed")
#                     current_progress["current_restaurant"] = rest_num
#                     scraped_current_progress["current_restaurant"] = rest_num
#                     self.save_current_progress()
#                     self.save_scraped_progress()
#                     self.commit_progress(f"Skipped restaurant {restaurant_name} on page {page_num} in {area_name} (already processed)")
#                     continue
                
#                 current_progress["current_restaurant"] = rest_num
#                 scraped_current_progress["current_restaurant"] = rest_num
                
#                 if any(category in restaurant['cuisine'] for category in skip_categories):
#                     print(f"\nSkipping restaurant {rest_num}/{len(restaurants_on_page)}: {restaurant_name} - Category: {restaurant['cuisine']}")
#                     if restaurant_name and restaurant_name not in current_progress["processed_restaurants"]:
#                         current_progress["processed_restaurants"].append(restaurant_name)
#                         scraped_current_progress["processed_restaurants"].append(restaurant_name)
#                     self.save_current_progress()
#                     self.save_scraped_progress()
#                     self.commit_progress(f"Skipped restaurant {restaurant_name} on page {page_num} in {area_name}")
#                     continue
                
#                 print(f"\nProcessing restaurant {rest_num}/{len(restaurants_on_page)} on page {page_num}: {restaurant_name}")
                
#                 try:
#                     restaurant.setdefault("menu_items", {})
#                     restaurant.setdefault("info", {})
#                     restaurant.setdefault("reviews", {})
#                     restaurant["page"] = page_num
                    
#                     # Set a timeout for scraping tasks (60 seconds per task)
#                     async def timeout_task(task, timeout=60):
#                         try:
#                             return await asyncio.wait_for(task, timeout=timeout)
#                         except asyncio.TimeoutError:
#                             print(f"Timeout while processing task for {restaurant_name}")
#                             logging.error(f"Timeout while processing task for {restaurant_name}")
#                             return None
                    
#                     print(f"Fetching menu for {restaurant_name}...")
#                     menu_data = await timeout_task(self.talabat_scraper.get_restaurant_menu(restaurant['url']))
#                     if menu_data:
#                         restaurant['menu_items'] = menu_data
#                     else:
#                         print(f"No menu data retrieved for {restaurant_name}")
                    
#                     print(f"Fetching info for {restaurant_name}...")
#                     info_data = await timeout_task(self.talabat_scraper.get_restaurant_info(restaurant['url']))
#                     if info_data:
#                         restaurant['info'] = info_data
#                     else:
#                         print(f"No info data retrieved for {restaurant_name}")
                    
#                     if restaurant['info'].get('Reviews URL') and restaurant['info']['Reviews URL'] != 'Not Available':
#                         print(f"Fetching reviews for {restaurant_name}...")
#                         reviews_data = self.talabat_scraper.get_reviews_data(restaurant['info']['Reviews URL'])
#                         restaurant['reviews'] = reviews_data or {}
#                     else:
#                         print(f"No reviews URL available for {restaurant_name}")
                    
#                     # Append to lists after all data is collected
#                     page_restaurants.append(restaurant)
#                     all_area_results.append(restaurant)
#                     if restaurant_name and restaurant_name not in current_progress["processed_restaurants"]:
#                         current_progress["processed_restaurants"].append(restaurant_name)
#                         scraped_current_progress["processed_restaurants"].append(restaurant_name)
#                     self.scraped_progress["all_results"][area_name] = all_area_results
#                     logging.debug(f"Updated all_results for {area_name}: {len(all_area_results)} restaurants")
                    
#                     # Save JSON and progress after each restaurant
#                     try:
#                         json_filename = os.path.join(self.output_dir, f"{area_name}.json")
#                         with open(json_filename, 'w', encoding='utf-8') as f:
#                             json.dump(all_area_results, f, indent=2, ensure_ascii=False)
#                         logging.info(f"Saved {len(all_area_results)} restaurants to {json_filename}")
                        
#                         # Log progress state before saving
#                         logging.debug(f"Current progress before save: {json.dumps(self.current_progress, ensure_ascii=False)}")
#                         logging.debug(f"Scraped progress before save: {json.dumps(self.scraped_progress, ensure_ascii=False)}")
                        
#                         self.save_current_progress()
#                         self.save_scraped_progress()
#                         self.print_progress_details()
#                         self.commit_progress(f"Processed restaurant {restaurant_name} on page {page_num} in {area_name}")
                    
#                     except Exception as e:
#                         print(f"Failed to save JSON or progress for {restaurant_name}: {e}")
#                         logging.error(f"Failed to save JSON or progress for {restaurant_name}: {e}")
                    
#                     await asyncio.sleep(2)
                
#                 except Exception as e:
#                     print(f"Error processing restaurant {rest_num}/{len(restaurants_on_page)}: {restaurant_name}: {e}")
#                     logging.error(f"Error processing restaurant {restaurant_name}: {e}")
#                     import traceback
#                     traceback.print_exc()
#                     if restaurant_name and restaurant_name not in current_progress["processed_restaurants"]:
#                         current_progress["processed_restaurants"].append(restaurant_name)
#                         scraped_current_progress["processed_restaurants"].append(restaurant_name)
#                     self.save_current_progress()
#                     self.save_scraped_progress()
#                     self.commit_progress(f"Error processing restaurant {restaurant_name} on page {page_num} in {area_name}")
            
#             # Save restaurants for the page to detailed Excel
#             if page_restaurants:
#                 try:
#                     print(f"Saving {len(page_restaurants)} restaurants from page {page_num} to {detailed_excel_filename} in sheet {area_name}")
#                     self.create_detailed_excel_sheet(area_name, page_restaurants, detailed_excel_filename)
#                 except Exception as e:
#                     print(f"Failed to save detailed Excel for page {page_num}: {e}")
#                     logging.error(f"Failed to save detailed Excel for page {page_num}: {e}")
            
#             # Mark page as complete
#             if page_num not in current_progress["completed_pages"]:
#                 current_progress["completed_pages"].append(page_num)
#                 scraped_current_progress["completed_pages"].append(page_num)
#             current_progress["current_restaurant"] = 0
#             scraped_current_progress["current_restaurant"] = 0
#             self.save_current_progress()
#             self.save_scraped_progress()
#             self.commit_progress(f"Completed page {page_num} in {area_name}")
#             await asyncio.sleep(3)
        
#         # Final JSON save
#         json_filename = os.path.join(self.output_dir, f"{area_name}.json")
#         with open(json_filename, 'w', encoding='utf-8') as f:
#             json.dump(all_area_results, f, indent=2, ensure_ascii=False)
#         logging.info(f"Final save: {len(all_area_results)} restaurants to {json_filename}")
        
#         # Create simplified Excel workbook
#         simplified_workbook = Workbook()
#         self.create_excel_sheet(simplified_workbook, area_name, all_area_results)
#         simplified_excel_filename = os.path.join(self.output_dir, f"{area_name}.xlsx")
#         simplified_workbook.save(simplified_excel_filename)
#         print(f"Simplified Excel file saved: {simplified_excel_filename}")
        
#         # Upload both Excel files to Google Drive
#         if self.upload_to_drive(simplified_excel_filename):
#             print(f"Uploaded {simplified_excel_filename} to Google Drive")
#         else:
#             print(f"Failed to upload {simplified_excel_filename} to Google Drive")
        
#         if self.upload_to_drive(detailed_excel_filename):
#             print(f"Uploaded {detailed_excel_filename} to Google Drive")
#         else:
#             print(f"Failed to upload {detailed_excel_filename} to Google Drive")
        
#         current_progress.update({
#             "area_name": None,
#             "current_page": 0,
#             "total_pages": 0,
#             "current_restaurant": 0,
#             "total_restaurants": 0,
#             "processed_restaurants": [],
#             "completed_pages": []
#         })
#         scraped_current_progress.update(current_progress)
#         self.save_current_progress()
#         self.save_scraped_progress()
#         self.print_progress_details()
#         self.commit_progress(f"Completed area {area_name}")
        
#         print(f"Saved {len(all_area_results)} restaurants for {area_name}")
#         return all_area_results
    
#     def save_scraped_progress(self, progress: Dict = None):
#         if progress is None:
#             progress = self.scraped_progress
#         try:
#             progress["last_updated"] = datetime.now().isoformat()
#             if "current_progress" in progress:
#                 progress["current_progress"]["completed_pages"] = sorted(list(set(
#                     int(page) for page in progress["current_progress"].get("completed_pages", [])
#                     if isinstance(page, (int, float)) and page >= 1
#                 )))
#             # Log content before saving
#             content_str = json.dumps(progress, ensure_ascii=False)
#             logging.debug(f"Saving scraped_progress content: {content_str}")
#             json.dumps(progress, ensure_ascii=False)  # Validate serializability
#             with tempfile.NamedTemporaryFile('w', delete=False, dir='.') as temp_file:
#                 json.dump(progress, temp_file, indent=2, ensure_ascii=False)
#                 temp_file.flush()
#                 os.fsync(temp_file.fileno())
#                 temp_filename = temp_file.name
#             os.replace(temp_filename, self.SCRAPED_PROGRESS_FILE)
#             # Verify file content after write
#             with open(self.SCRAPED_PROGRESS_FILE, 'r', encoding='utf-8') as f:
#                 written_content = f.read()
#             logging.debug(f"Verified scraped_progress.json content after write: {written_content}")
#             mtime = os.path.getmtime(self.SCRAPED_PROGRESS_FILE)
#             print(f"Saved scraped progress to {self.SCRAPED_PROGRESS_FILE} at {datetime.fromtimestamp(mtime).isoformat()}")
#             logging.info(f"Saved scraped progress to {self.SCRAPED_PROGRESS_FILE}")
#         except Exception as e:
#             print(f"Failed to save scraped progress: {e}")
#             logging.error(f"Failed to save scraped progress: {e}")
#             raise
    
#     def commit_progress(self, message: str):
#         try:
#             # Check git status before staging
#             status_result = subprocess.run(["git", "status", "--porcelain"], capture_output=True, text=True, check=True)
#             logging.debug(f"Git status before staging: {status_result.stdout}")
            
#             # Explicitly stage files
#             subprocess.run(["git", "add", self.CURRENT_PROGRESS_FILE], check=True)
#             subprocess.run(["git", "add", self.SCRAPED_PROGRESS_FILE], check=True)
#             subprocess.run(["git", "add", self.output_dir], check=True)
            
#             # Check git diff to confirm changes
#             diff_result = subprocess.run(["git", "diff", "--staged"], capture_output=True, text=True, check=True)
#             logging.debug(f"Git diff --staged: {diff_result.stdout}")
            
#             # Attempt to commit
#             result = subprocess.run(["git", "commit", "-m", message], capture_output=True, text=True)
#             if result.returncode == 0:
#                 print(f"Committed progress: {message}")
#                 logging.info(f"Committed progress: {message}")
#             else:
#                 print(f"No changes to commit for: {message}")
#                 logging.warning(f"No changes to commit: {result.stderr}")
            
#             # Push to remote
#             push_result = subprocess.run(["git", "push"], capture_output=True, text=True)
#             if push_result.returncode == 0:
#                 print(f"Pushed progress: {message}")
#                 logging.info(f"Pushed progress: {message}")
#             else:
#                 print(f"Failed to push progress: {push_result.stderr}")
#                 logging.error(f"Failed to push progress: {push_result.stderr}")
            
#         except subprocess.CalledProcessError as e:
#             print(f"Failed to commit progress: {e}")
#             logging.error(f"Failed to commit progress: {e}")
    
#     async def determine_total_pages(self, area_url: str) -> int:
#         print(f"Determining total pages for URL: {area_url}")
#         try:
#             async with async_playwright() as p:
#                 browser = await p.firefox.launch(headless=True)
#                 context = await browser.new_context(
#                     viewport={'width': 1920, 'height': 1080},
#                     user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
#                 )
#                 page = await context.new_page()
#                 page.set_default_timeout(120000)
                
#                 response = await page.goto(area_url, wait_until='domcontentloaded')
#                 if not response or not response.ok:
#                     print(f"Failed to load page: {response.status if response else 'No response'}")
#                     return 1
                
#                 await page.wait_for_selector("ul[data-test='pagination'], .vendor-card, [data-testid='restaurant-a']", timeout=30000)
                
#                 last_page = 1
#                 pagination = await page.query_selector("ul[data-test='pagination']")
#                 if pagination:
#                     items = await pagination.query_selector_all("li[data-testid='paginate-link']")
#                     if items and len(items) > 1:
#                         last_page_item = items[-2]
#                         last_page_link = await last_page_item.query_selector("a[page]")
#                         if last_page_link:
#                             last_page_attr = await last_page_link.get_attribute("page")
#                             if last_page_attr and last_page_attr.isdigit():
#                                 last_page = int(last_page_attr)
                
#                 await browser.close()
#                 return last_page
#         except Exception as e:
#             print(f"Error determining total pages: {e}")
#             return 1

#     async def get_page_restaurants(self, page_url: str, page_num: int) -> List[Dict]:
#         browser = None
#         try:
#             async with async_playwright() as p:
#                 browser = await p.firefox.launch(headless=True)
#                 context = await browser.new_context(
#                     viewport={'width': 1920, 'height': 1080},
#                     user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
#                 )
#                 page = await context.new_page()
#                 page.set_default_timeout(120000)
                
#                 response = await page.goto(page_url, wait_until='domcontentloaded')
#                 if not response or not response.ok:
#                     print(f"Failed to load page {page_num}: {response.status if response else 'No response'}")
#                     return []
                
#                 await page.wait_for_selector(".vendor-card, [data-testid='restaurant-a']", timeout=30000)
#                 return await self.talabat_scraper._extract_restaurants_from_page(page, page_num)
#         except Exception as e:
#             print(f"Error getting page restaurants: {e}")
#             import traceback
#             traceback.print_exc()
#             return []
#         finally:
#             if browser:
#                 await browser.close()

#     def create_excel_sheet(self, workbook, sheet_name: str, data: List[Dict]):
#         sheet = workbook.create_sheet(title=sheet_name)
#         try:
#             simplified_data = []
#             for restaurant in data:
#                 restaurant_info = {
#                     "Name": restaurant.get("name", ""),
#                     "Cuisine": restaurant.get("cuisine", ""),
#                     "Rating": restaurant.get("rating", ""),
#                     "Delivery Time": restaurant.get("delivery_time", ""),
#                     "Delivery Fee": restaurant.get("delivery_fee", ""),
#                     "Min Order": restaurant.get("min_order", ""),
#                     "URL": restaurant.get("url", ""),
#                 }
#                 if restaurant.get("info"):
#                     restaurant_info.update({
#                         "Address": restaurant["info"].get("Address", ""),
#                         "Working Hours": restaurant["info"].get("Working Hours", ""),
#                     })
#                 if restaurant.get("reviews") and restaurant["reviews"].get("Rating_value"):
#                     restaurant_info.update({
#                         "Rating Value": restaurant["reviews"]["Rating_value"],
#                         "Ratings Count": restaurant["reviews"].get("Ratings_count", ""),
#                         "Reviews Count": restaurant["reviews"].get("Reviews_count", ""),
#                     })
#                 if restaurant.get("menu_items"):
#                     restaurant_info["Menu Categories"] = len(restaurant["menu_items"])
#                     item_count = sum(len(items) for items in restaurant["menu_items"].values())
#                     restaurant_info["Menu Items"] = item_count
#                 simplified_data.append(restaurant_info)
            
#             if simplified_data:
#                 df = pd.DataFrame(simplified_data)
#                 for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
#                     for c_idx, value in enumerate(row, 1):
#                         sheet.cell(row=r_idx, column=c_idx, value=value)
#                 for column in sheet.columns:
#                     max_length = max(len(str(cell.value or "")) for cell in column)
#                     column_letter = get_column_letter(column[0].column)
#                     sheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
#             else:
#                 sheet.cell(row=1, column=1, value="No data found for this area")
#         except Exception as e:
#             print(f"Error creating Excel sheet for {sheet_name}: {str(e)}")
#             sheet.cell(row=1, column=1, value=f"Error processing data: {str(e)}")

#     def flatten_menu_items(self, menu_items):
#         if not isinstance(menu_items, dict):
#             return ''
#         items = []
#         for category, dishes in menu_items.items():
#             for dish in dishes:
#                 if isinstance(dish, dict):
#                     name = dish.get('name', '')
#                     price = dish.get('prices', {}).get('new_price', '')
#                     items.append(f"{name}: {price}")
#         return '; '.join(items)

#     def create_detailed_excel_sheet(self, area_name: str, data: List[Dict], excel_filename: str):
#         columns = [
#             'restaurant name', 'cuisine', 'restaurant url', 'general rating',
#             'restaurant in page number', 'delivery time', 'delivery fees',
#             'minimum order', 'tracking status', 'contactless', 'menu items',
#             'Address', 'Reviews URL', 'Pre-Order status', 'Payment types',
#             'reviews rating value', 'reviews Ratings count', 'reviews count',
#             'General review', 'Order Packaging reviews', 'Value for money reviews',
#             'Delivery time reviews', 'Quality of food reviews', 'Customer reviews'
#         ]
        
#         try:
#             # Prepare data for DataFrame
#             rows = []
#             for restaurant in data:
#                 row = {
#                     'restaurant name': restaurant.get('name', ''),
#                     'cuisine': restaurant.get('cuisine', ''),
#                     'restaurant url': restaurant.get('url', ''),
#                     'general rating': restaurant.get('rating', ''),
#                     'restaurant in page number': restaurant.get('page', ''),
#                     'delivery time': restaurant.get('delivery_time', ''),
#                     'delivery fees': restaurant.get('delivery_fee', ''),
#                     'minimum order': restaurant.get('min_order', ''),
#                     'tracking status': restaurant.get('tracking_status', ''),
#                     'contactless': restaurant.get('contactless', ''),
#                     'menu items': self.flatten_menu_items(restaurant.get('menu_items', {})),
#                     'Address': restaurant.get('info', {}).get('Address', ''),
#                     'Reviews URL': restaurant.get('info', {}).get('Reviews URL', ''),
#                     'Pre-Order status': restaurant.get('info', {}).get('Pre-Order', ''),
#                     'Payment types': ', '.join(restaurant.get('info', {}).get('Payment', [])),
#                     'reviews rating value': restaurant.get('reviews', {}).get('Rating_value', ''),
#                     'reviews Ratings count': restaurant.get('reviews', {}).get('Ratings_count', ''),
#                     'reviews count': restaurant.get('reviews', {}).get('Reviews_count', ''),
#                     'General review': '; '.join(restaurant.get('reviews', {}).get('General_review', [])),
#                     'Order Packaging reviews': restaurant.get('reviews', {}).get('Specific_reviews', {}).get('Order Packaging', ''),
#                     'Value for money reviews': restaurant.get('reviews', {}).get('Specific_reviews', {}).get('Value for money', ''),
#                     'Delivery time reviews': restaurant.get('reviews', {}).get('Specific_reviews', {}).get('Delivery time', ''),
#                     'Quality of food reviews': restaurant.get('reviews', {}).get('Specific_reviews', {}).get('Quality of food', ''),
#                     'Customer reviews': '; '.join([
#                         f"{rev.get('reviewer_name', 'Unknown')}: {rev.get('review_comment', '')} ({rev.get('review_date', '')})"
#                         for rev in restaurant.get('reviews', {}).get('Customer_reviews', [])
#                         if isinstance(rev, dict)
#                     ])
#                 }
#                 rows.append(row)
            
#             if not rows:
#                 print(f"No data to save for {area_name}")
#                 return
            
#             # Create DataFrame for new data
#             new_df = pd.DataFrame(rows, columns=columns)
            
#             # Check if Excel file exists
#             if os.path.exists(excel_filename):
#                 workbook = load_workbook(excel_filename)
#                 if area_name in workbook.sheetnames:
#                     # Read existing sheet data
#                     existing_df = pd.read_excel(excel_filename, sheet_name=area_name)
#                     # Append new data
#                     combined_df = pd.concat([existing_df, new_df], ignore_index=True)
#                     # Remove existing sheet to update with combined data
#                     del workbook[area_name]
#                 else:
#                     combined_df = new_df
#             else:
#                 workbook = Workbook()
#                 if "Sheet" in workbook.sheetnames:
#                     workbook.remove(workbook["Sheet"])
#                 combined_df = new_df
            
#             # Create or update sheet
#             sheet = workbook.create_sheet(title=area_name)
#             for r_idx, row in enumerate(dataframe_to_rows(combined_df, index=False, header=True), 1):
#                 for c_idx, value in enumerate(row, 1):
#                     sheet.cell(row=r_idx, column=c_idx, value=value)
            
#             # Adjust column widths
#             for column in sheet.columns:
#                 max_length = max(len(str(cell.value or "")) for cell in column)
#                 column_letter = get_column_letter(column[0].column)
#                 sheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
            
#             # Save workbook
#             workbook.save(excel_filename)
#             print(f"Saved {len(rows)} restaurants to {excel_filename} in sheet {area_name}")
        
#         except Exception as e:
#             print(f"Error saving detailed Excel sheet for {area_name}: {str(e)}")
#             logging.error(f"Error saving detailed Excel sheet for {area_name}: {str(e)}")

#     @retry(tries=3, delay=2, backoff=2)
#     def upload_to_drive(self, file_path):
#         print(f"\nUploading {file_path} to Google Drive...")
#         try:
#             credentials_json = os.environ.get('TALABAT_GCLOUD_KEY_JSON')
#             if not credentials_json:
#                 print("Error: TALABAT_GCLOUD_KEY_JSON environment variable is empty or not set!")
#                 return False
#             self.drive_uploader = SavingOnDrive(credentials_json=credentials_json)
#             if not self.drive_uploader.authenticate():
#                 print("Failed to authenticate with Google Drive. Check TALABAT_GCLOUD_KEY_JSON validity.")
#                 return False
#             file_ids = self.drive_uploader.upload_to_multiple_folders(file_path)
#             success = len(file_ids) == 2
#             if success:
#                 print(f"Successfully uploaded {file_path} to Google Drive")
#             else:
#                 print(f"Failed to upload {file_path}: Incomplete upload to folders")
#             return success
#         except Exception as e:
#             print(f"Error uploading to Google Drive: {str(e)}")
#             return False

#     async def run(self):
#         ahmadi_areas = [
#             ("الظهر", "https://www.talabat.com/kuwait/restaurants/59/dhaher"),
#             ("الرقه", "https://www.talabat.com/kuwait/restaurants/37/riqqa"),
#             ("هدية", "https://www.talabat.com/kuwait/restaurants/30/hadiya"),
#             ("المنقف", "https://www.talabat.com/kuwait/restaurants/32/mangaf"),
#             ("أبو حليفة", "https://www.talabat.com/kuwait/restaurants/2/abu-halifa"),
#             ("الفنطاس", "https://www.talabat.com/kuwait/restaurants/38/fintas"),
#             ("العقيلة", "https://www.talabat.com/kuwait/restaurants/79/egaila"),
#             ("الصباحية", "https://www.talabat.com/kuwait/restaurants/31/sabahiya"),
#             ("الأحمدي", "https://www.talabat.com/kuwait/restaurants/3/al-ahmadi"),
#             ("الفحيحيل", "https://www.talabat.com/kuwait/restaurants/5/fahaheel"),
#             ("شرق الأحمدي", "https://www.talabat.com/kuwait/restaurants/3/al-ahmadi"),
#             ("ضاحية علي صباح السالم", "https://www.talabat.com/kuwait/restaurants/82/ali-sabah-al-salem-umm-al-hayman"),
#             ("ميناء عبد الله", "https://www.talabat.com/kuwait/restaurants/100/mina-abdullah"),
#             ("بنيدر", "https://www.talabat.com/kuwait/restaurants/6650/bnaider"),
#             ("الزور", "https://www.talabat.com/kuwait/restaurants/2053/zour"),
#             ("الجليعة", "https://www.talabat.com/kuwait/restaurants/6860/al-julaiaa"),
#             ("المهبولة", "https://www.talabat.com/kuwait/restaurants/24/mahboula"),
#             ("النويصيب", "https://www.talabat.com/kuwait/restaurants/2054/nuwaiseeb"),
#             ("الخيران", "https://www.talabat.com/kuwait/restaurants/2726/khairan"),
#             ("الوفرة", "https://www.talabat.com/kuwait/restaurants/2057/wafra-farms"),
#             ("ضاحية فهد الأحمد", "https://www.talabat.com/kuwait/restaurants/98/fahad-al-ahmed"),
#             ("ضاحية جابر العلي", "https://www.talabat.com/kuwait/restaurants/60/jaber-al-ali"),
#             ("مدينة صباح الأحمد السكنية", "https://www.talabat.com/kuwait/restaurants/6931/sabah-al-ahmad-2"),
#             ("مدينة صباح الأحمد البحرية", "https://www.talabat.com/kuwait/restaurants/2726/khairan"),
#             ("ميناء الأحمدي", "https://www.talabat.com/kuwait/restaurants/3/al-ahmadi")
#         ]
        
#         simplified_excel_filename = os.path.join(self.output_dir, "الاحمدي.xlsx")
#         simplified_workbook = Workbook()
#         if "Sheet" in simplified_workbook.sheetnames:
#             simplified_workbook.remove(simplified_workbook["Sheet"])
        
#         completed_areas = self.current_progress["completed_areas"]
#         current_area_index = self.current_progress["current_area_index"]
        
#         print(f"Starting from area index {current_area_index}")
#         print(f"Already completed areas: {', '.join(completed_areas) if completed_areas else 'None'}")
        
#         resuming_area = self.current_progress["current_progress"]["area_name"]
#         if resuming_area:
#             for idx, (area_name, _) in enumerate(ahmadi_areas):
#                 if area_name == resuming_area:
#                     print(f"Resuming from area {resuming_area} (index {idx})")
#                     current_area_index = idx
#                     self.current_progress["current_area_index"] = idx
#                     self.scraped_progress["current_area_index"] = idx
#                     self.save_current_progress()
#                     self.save_scraped_progress()
#                     self.commit_progress(f"Resuming from area {resuming_area}")
#                     break
        
#         for idx, (area_name, area_url) in enumerate(ahmadi_areas):
#             if area_name in completed_areas and area_name != resuming_area:
#                 print(f"Skipping completed area: {area_name}")
#                 continue
#             if idx < current_area_index:
#                 print(f"Skipping area {area_name} (index {idx} < {current_area_index})")
#                 continue
            
#             self.current_progress["current_area_index"] = idx
#             self.scraped_progress["current_area_index"] = idx
#             self.save_current_progress()
#             self.save_scraped_progress()
#             self.commit_progress(f"Starting area {area_name} at index {idx}")
            
#             try:
#                 area_results = await self.scrape_and_save_area(area_name, area_url)
#                 self.create_excel_sheet(simplified_workbook, area_name, area_results)
#                 simplified_workbook.save(simplified_excel_filename)
#                 print(f"Updated simplified Excel file: {simplified_excel_filename}")
                
#                 if area_name not in completed_areas:
#                     completed_areas.append(area_name)
#                     self.current_progress["completed_areas"] = completed_areas
#                     self.scraped_progress["completed_areas"] = completed_areas
#                 self.save_current_progress()
#                 self.save_scraped_progress()
#                 self.print_progress_details()
#                 self.commit_progress(f"Completed area {area_name} in run")
#                 await asyncio.sleep(5)
            
#             except Exception as e:
#                 print(f"Error processing area {area_name}: {e}")
#                 logging.error(f"Error processing area {area_name}: {e}")
#                 import traceback
#                 traceback.print_exc()
#                 self.save_current_progress()
#                 self.save_scraped_progress()
#                 self.commit_progress(f"Progress update after error in {area_name}")
        
#         simplified_workbook.save(simplified_excel_filename)
#         combined_json_filename = os.path.join(self.output_dir, "الاحمدي_all.json")
#         with open(combined_json_filename, 'w', encoding='utf-8') as f:
#             json.dump(self.scraped_progress["all_results"], f, indent=2, ensure_ascii=False)
        
#         print(f"\n{'='*50}")
#         print(f"SCRAPING COMPLETED")
#         print(f"Simplified Excel file saved: {simplified_excel_filename}")
#         print(f"Combined JSON saved: {combined_json_filename}")
        
#         if len(completed_areas) == len(ahmadi_areas):
#             if self.upload_to_drive(simplified_excel_filename):
#                 print(f"Uploaded simplified Excel file to Google Drive")
#             else:
#                 print(f"Failed to upload simplified Excel file to Google Drive")
#         else:
#             print(f"Scraping incomplete ({len(completed_areas)}/{len(ahmadi_areas)} areas)")
        
#         self.commit_progress("Final progress update after run")

# async def main():
#     try:
#         scraper = MainScraper()
#         await scraper.run()
#     except KeyboardInterrupt:
#         print("\nInterrupted. Saving progress...")
#         if 'scraper' in locals():
#             scraper.save_current_progress()
#             scraper.save_scraped_progress()
#             scraper.commit_progress("Progress saved after interruption")
#         print("Progress saved. Exiting.")
#     except Exception as e:
#         print(f"Critical error: {e}")
#         import traceback
#         traceback.print_exc()
#         if 'scraper' in locals():
#             scraper.save_current_progress()
#             scraper.save_scraped_progress()
#             scraper.commit_progress("Progress saved after critical error")
#         sys.exit(1)
        
# if __name__ == "__main__":
#     scraper = MainScraper()
#     scraper.print_progress_details()
#     asyncio.run(scraper.run())




# import asyncio
# import json
# import os
# import tempfile
# import sys
# import subprocess
# from retry import retry
# import re
# from typing import Dict, List, Tuple
# import pandas as pd
# from openpyxl import Workbook
# from openpyxl.utils.dataframe import dataframe_to_rows
# from openpyxl.utils import get_column_letter
# from playwright.async_api import async_playwright
# from talabat_main_scraper import TalabatScraper
# from SavingOnDrive import SavingOnDrive
# from time import sleep
# from datetime import datetime
# import logging

# logging.basicConfig(
#     filename='scraper.log',
#     level=logging.DEBUG,
#     format='%(asctime)s - %(levelname)s - %(message)s'
# )

# class MainScraper:
#     CURRENT_PROGRESS_FILE = "current_progress.json"
#     SCRAPED_PROGRESS_FILE = "scraped_progress.json"

#     def __init__(self):
#         self.talabat_scraper = TalabatScraper()
#         self.output_dir = "output"
#         credentials_json = os.environ.get('TALABAT_GCLOUD_KEY_JSON')
#         self.drive_uploader = SavingOnDrive(credentials_json=credentials_json)
        
#         os.makedirs(self.output_dir, exist_ok=True)
        
#         self.current_progress = self.load_current_progress()
#         self.scraped_progress = self.load_scraped_progress()
        
#         self.github_token = os.environ.get('GITHUB_TOKEN')
#         self.ensure_playwright_browsers()

#     def ensure_playwright_browsers(self):
#         try:
#             print("Installing Playwright browsers...")
#             subprocess.run([sys.executable, "-m", "playwright", "install", "chromium", "firefox"], 
#                           check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
#             print("Playwright browsers installed successfully")
#         except subprocess.CalledProcessError as e:
#             print(f"Error installing Playwright browsers: {e}")
#             logging.error(f"Error installing Playwright browsers: {e}")

#     # def load_current_progress(self) -> Dict:
#     #     if os.path.exists(self.CURRENT_PROGRESS_FILE):
#     #         try:
#     #             with open(self.CURRENT_PROGRESS_FILE, 'r', encoding='utf-8') as f:
#     #                 progress = json.load(f)
#     #             print(f"Loaded current progress from {self.CURRENT_PROGRESS_FILE}")
#     #             return progress
#     #         except Exception as e:
#     #             print(f"Error loading current progress: {e}")
        
#     #     default_progress = {
#     #         "completed_areas": [],
#     #         "current_area_index": 0,
#     #         "last_updated": None,
#     #         "current_progress": {
#     #             "area_name": None,
#     #             "current_page": 0,
#     #             "total_pages": 0,
#     #             "current_restaurant": 0,  # 0 means no restaurant processed yet
#     #             "total_restaurants": 0,
#     #             "processed_restaurants": [],
#     #             "completed_pages": []
#     #         }
#     #     }
#     #     self.save_current_progress(default_progress)
#     #     return default_progress

#     def load_current_progress(self) -> Dict:
#         default_progress = {
#             "completed_areas": [],
#             "current_area_index": 0,
#             "last_updated": None,
#             "current_progress": {
#                 "area_name": None,
#                 "current_page": 0,
#                 "total_pages": 0,
#                 "current_restaurant": 0,
#                 "total_restaurants": 0,
#                 "processed_restaurants": [],
#                 "completed_pages": []
#             }
#         }
#         if not os.path.exists(self.CURRENT_PROGRESS_FILE):
#             print(f"No current progress file found, initializing {self.CURRENT_PROGRESS_FILE}")
#             self.save_current_progress(default_progress)
#             return default_progress
        
#         try:
#             with open(self.CURRENT_PROGRESS_FILE, 'r', encoding='utf-8') as f:
#                 progress = json.load(f)
#             # Validate structure
#             if not isinstance(progress, dict) or "current_progress" not in progress:
#                 print(f"Invalid current progress file, resetting to default")
#                 logging.warning(f"Invalid current progress file structure")
#                 self.save_current_progress(default_progress)
#                 return default_progress
#             # Deduplicate processed_restaurants and completed_pages
#             progress["current_progress"]["processed_restaurants"] = list(set(
#                 str(item) for item in progress["current_progress"].get("processed_restaurants", [])
#             ))
#             progress["current_progress"]["completed_pages"] = sorted(list(set(
#                 int(page) for page in progress["current_progress"].get("completed_pages", [])
#                 if isinstance(page, (int, float)) and page >= 1
#             )))
#             print(f"Loaded current progress from {self.CURRENT_PROGRESS_FILE}")
#             logging.info(f"Loaded current progress: {json.dumps(progress, ensure_ascii=False)}")
#             return progress
#         except Exception as e:
#             print(f"Error loading current progress: {e}")
#             logging.error(f"Error loading current progress: {e}")
#             self.save_current_progress(default_progress)
#             return default_progress

#     # def save_current_progress(self, progress: Dict = None):
#     #     if progress is None:
#     #         progress = self.current_progress
#     #     try:
#     #         progress["last_updated"] = datetime.now().isoformat()
#     #         with tempfile.NamedTemporaryFile('w', delete=False, dir='.') as temp_file:
#     #             json.dump(progress, temp_file, indent=2, ensure_ascii=False)
#     #             temp_file.flush()
#     #             os.fsync(temp_file.fileno())
#     #             temp_filename = temp_file.name
#     #         os.replace(temp_filename, self.CURRENT_PROGRESS_FILE)
#     #         print(f"Saved current progress to {self.CURRENT_PROGRESS_FILE}")
#     #     except Exception as e:
#     #         print(f"Error saving current progress: {e}")

#     def save_current_progress(self, progress: Dict = None):
#         if progress is None:
#             progress = self.current_progress
#         try:
#             progress["last_updated"] = datetime.now().isoformat()
#             # Deduplicate processed_restaurants and completed_pages
#             if "current_progress" in progress:
#                 progress["current_progress"]["processed_restaurants"] = list(set(
#                     str(item) for item in progress["current_progress"].get("processed_restaurants", [])
#                 ))
#                 progress["current_progress"]["completed_pages"] = sorted(list(set(
#                     int(page) for page in progress["current_progress"].get("completed_pages", [])
#                     if isinstance(page, (int, float)) and page >= 1
#                 )))
#             # Validate JSON serializability
#             json.dumps(progress, ensure_ascii=False)
#             with tempfile.NamedTemporaryFile('w', delete=False, dir='.') as temp_file:
#                 json.dump(progress, temp_file, indent=2, ensure_ascii=False)
#                 temp_file.flush()
#                 os.fsync(temp_file.fileno())
#                 temp_filename = temp_file.name
#             os.replace(temp_filename, self.CURRENT_PROGRESS_FILE)
#             print(f"Saved current progress to {self.CURRENT_PROGRESS_FILE}")
#             logging.info(f"Saved current progress: {json.dumps(progress, ensure_ascii=False)}")
#         except Exception as e:
#             print(f"Failed to save current progress: {e}")
#             logging.error(f"Failed to save current progress: {e}")

#     # def load_scraped_progress(self) -> Dict:
#     #     if os.path.exists(self.SCRAPED_PROGRESS_FILE):
#     #         try:
#     #             with open(self.SCRAPED_PROGRESS_FILE, 'r', encoding='utf-8') as f:
#     #                 progress = json.load(f)
#     #             print(f"Loaded scraped progress from {self.SCRAPED_PROGRESS_FILE}")
#     #             return progress
#     #         except Exception as e:
#     #             print(f"Error loading scraped progress: {e}")
        
#     #     default_progress = {
#     #         "completed_areas": [],
#     #         "current_area_index": 0,
#     #         "last_updated": None,
#     #         "all_results": {},
#     #         "current_progress": {
#     #             "area_name": None,
#     #             "current_page": 0,
#     #             "total_pages": 0,
#     #             "current_restaurant": 0,  # 0 means no restaurant processed yet
#     #             "total_restaurants": 0,
#     #             "processed_restaurants": [],
#     #             "completed_pages": []
#     #         }
#     #     }
#     #     self.save_scraped_progress(default_progress)
#     #     return default_progress

#     def load_scraped_progress(self) -> Dict:
#         default_progress = {
#             "completed_areas": [],
#             "current_area_index": 0,
#             "last_updated": None,
#             "all_results": {},
#             "current_progress": {
#                 "area_name": None,
#                 "current_page": 0,
#                 "total_pages": 0,
#                 "current_restaurant": 0,
#                 "total_restaurants": 0,
#                 "processed_restaurants": [],
#                 "completed_pages": []
#             }
#         }
#         if not os.path.exists(self.SCRAPED_PROGRESS_FILE):
#             print(f"No scraped progress file found, initializing {self.SCRAPED_PROGRESS_FILE}")
#             self.save_scraped_progress(default_progress)
#             return default_progress
        
#         try:
#             with open(self.SCRAPED_PROGRESS_FILE, 'r', encoding='utf-8') as f:
#                 progress = json.load(f)
#             # Validate structure
#             if not isinstance(progress, dict) or "current_progress" not in progress or "all_results" not in progress:
#                 print(f"Invalid scraped progress file, resetting to default")
#                 logging.warning(f"Invalid scraped progress file structure")
#                 self.save_scraped_progress(default_progress)
#                 return default_progress
#             # Deduplicate processed_restaurants and completed_pages
#             progress["current_progress"]["processed_restaurants"] = list(set(
#                 str(item) for item in progress["current_progress"].get("processed_restaurants", [])
#             ))
#             progress["current_progress"]["completed_pages"] = sorted(list(set(
#                 int(page) for page in progress["current_progress"].get("completed_pages", [])
#                 if isinstance(page, (int, float)) and page >= 1
#             )))
#             print(f"Loaded scraped progress from {self.SCRAPED_PROGRESS_FILE}")
#             logging.info(f"Loaded scraped progress: {json.dumps(progress, ensure_ascii=False)}")
#             return progress
#         except Exception as e:
#             print(f"Error loading scraped progress: {e}")
#             logging.error(f"Error loading scraped progress: {e}")
#             self.save_scraped_progress(default_progress)
#             return default_progress

#     # def save_scraped_progress(self, progress: Dict = None):
#     #     if progress is None:
#     #         progress = self.scraped_progress
#     #     try:
#     #         progress["last_updated"] = datetime.now().isoformat()
#     #         with tempfile.NamedTemporaryFile('w', delete=False, dir='.') as temp_file:
#     #             json.dump(progress, temp_file, indent=2, ensure_ascii=False)
#     #             temp_file.flush()
#     #             os.fsync(temp_file.fileno())
#     #             temp_filename = temp_file.name
#     #         os.replace(temp_filename, self.SCRAPED_PROGRESS_FILE)
#     #         print(f"Saved scraped progress to {self.SCRAPED_PROGRESS_FILE}")
#     #     except Exception as e:
#     #         print(f"Error saving scraped progress: {e}")

#     def save_scraped_progress(self, progress: Dict = None):
#         if progress is None:
#             progress = self.scraped_progress
#         try:
#             progress["last_updated"] = datetime.now().isoformat()
#             # Deduplicate processed_restaurants and completed_pages
#             if "current_progress" in progress:
#                 progress["current_progress"]["processed_restaurants"] = list(set(
#                     str(item) for item in progress["current_progress"].get("processed_restaurants", [])
#                 ))
#                 progress["current_progress"]["completed_pages"] = sorted(list(set(
#                     int(page) for page in progress["current_progress"].get("completed_pages", [])
#                     if isinstance(page, (int, float)) and page >= 1
#                 )))
#             # Validate JSON serializability
#             json.dumps(progress, ensure_ascii=False)
#             with tempfile.NamedTemporaryFile('w', delete=False, dir='.') as temp_file:
#                 json.dump(progress, temp_file, indent=2, ensure_ascii=False)
#                 temp_file.flush()
#                 os.fsync(temp_file.fileno())
#                 temp_filename = temp_file.name
#             os.replace(temp_filename, self.SCRAPED_PROGRESS_FILE)
#             print(f"Saved scraped progress to {self.SCRAPED_PROGRESS_FILE}")
#             logging.info(f"Saved scraped progress: {json.dumps(progress, ensure_ascii=False)}")
#         except Exception as e:
#             print(f"Failed to save scraped progress: {e}")
#             logging.error(f"Failed to save scraped progress: {e}")

#     # def commit_progress(self, message: str = "Periodic progress update"):
#     #     if not self.github_token:
#     #         print("No GITHUB_TOKEN available, skipping commit")
#     #         return
        
#     #     try:
#     #         subprocess.run(["git", "config", "--global", "user.name", "GitHub Action"], check=True)
#     #         subprocess.run(["git", "config", "--global", "user.email", "action@github.com"], check=True)
#     #         subprocess.run(["git", "add", self.CURRENT_PROGRESS_FILE, self.SCRAPED_PROGRESS_FILE, self.output_dir], check=True)
#     #         result = subprocess.run(["git", "commit", "-m", message], capture_output=True, text=True)
#     #         if result.returncode == 0 or "nothing to commit" in result.stdout:
#     #             subprocess.run(["git", "push"], check=True, env={"GIT_AUTH_TOKEN": self.github_token})
#     #             print(f"Committed progress: {message}")
#     #         else:
#     #             print("No changes to commit")
#     #     except subprocess.CalledProcessError as e:
#     #         print(f"Error committing progress: {e}")

#     def commit_progress(self, message: str = "Periodic progress update"):
#         try:
#             subprocess.run(["git", "add", self.CURRENT_PROGRESS_FILE, self.SCRAPED_PROGRESS_FILE, self.output_dir], check=True)
#             result = subprocess.run(["git", "commit", "-m", message], capture_output=True, text=True)
#             if result.returncode == 0 or "nothing to commit" in result.stdout:
#                 if not self.github_token:
#                     print("No GITHUB_TOKEN available, saving locally")
#                     self.save_current_progress()
#                     self.save_scraped_progress()
#                     return
#                 subprocess.run(["git", "push"], check=True, env={"GIT_AUTH_TOKEN": self.github_token})
#                 print(f"Committed and pushed progress: {message}")
#             else:
#                 print("No changes to commit")
#         except subprocess.CalledProcessError as e:
#             print(f"Failed to commit progress: {e}")
#             logging.error(f"Failed to commit progress: {e}")
#             self.save_current_progress()
#             self.save_scraped_progress()
#         except Exception as e:
#             print(f"Unexpected error during commit: {e}")
#             logging.error(f"Unexpected error during commit: {e}")
#             self.save_current_progress()
#             self.save_scraped_progress()

#     def print_progress_details(self):
#         try:
#             with open(self.CURRENT_PROGRESS_FILE, 'r', encoding='utf-8') as f:
#                 current = json.load(f)
#             print("\nCurrent Progress:")
#             print(json.dumps(current, indent=2, ensure_ascii=False))
#         except Exception as e:
#             print(f"Error printing current progress: {e}")
#             logging.error(f"Error printing current progress: {e}")

#     # async def scrape_and_save_area(self, area_name: str, area_url: str) -> List[Dict]:
#     #     print(f"\n{'='*50}")
#     #     print(f"SCRAPING AREA: {area_name}")
#     #     print(f"URL: {area_url}")
#     #     print(f"{'='*50}\n")
        
#     #     all_area_results = self.scraped_progress["all_results"].get(area_name, [])
#     #     current_progress = self.current_progress["current_progress"]
#     #     scraped_current_progress = self.scraped_progress["current_progress"]
        
#     #     is_resuming = current_progress["area_name"] == area_name
#     #     start_page = current_progress["current_page"] if is_resuming else 1
#     #     start_restaurant = current_progress["current_restaurant"] if is_resuming else 0  # 0 means start from first
        
#     #     if is_resuming:
#     #         print(f"Resuming area {area_name} from page {start_page} restaurant {start_restaurant + 1 if start_restaurant > 0 else 1}")
#     #     else:
#     #         current_progress.update({
#     #             "area_name": area_name,
#     #             "current_page": start_page,
#     #             "total_pages": 0,
#     #             "current_restaurant": 0,
#     #             "total_restaurants": 0,
#     #             "processed_restaurants": [],
#     #             "completed_pages": []
#     #         })
#     #         scraped_current_progress.update(current_progress)
#     #         self.save_current_progress()
#     #         self.save_scraped_progress()
#     #         self.commit_progress(f"Started scraping area {area_name}")
        
#     #     skip_categories = {"Grocery, Convenience Store", "Pharmacy", "Flowers", "Electronics", "Grocery, Hypermarket"}
        
#     #     if current_progress["total_pages"] == 0:
#     #         total_pages = await self.determine_total_pages(area_url)
#     #         current_progress["total_pages"] = total_pages
#     #         scraped_current_progress["total_pages"] = total_pages
#     #         self.save_current_progress()
#     #         self.save_scraped_progress()
#     #         self.commit_progress(f"Determined {total_pages} pages for {area_name}")
#     #     else:
#     #         total_pages = current_progress["total_pages"]
        
#     #     print(f"Total pages for {area_name}: {total_pages}")
        
#     #     for page_num in range(start_page, total_pages + 1):
#     #         if page_num in current_progress["completed_pages"]:
#     #             print(f"Skipping completed page {page_num}")
#     #             continue
            
#     #         page_url = area_url if page_num == 1 else (
#     #             re.sub(r'page=\d+', f'page={page_num}', area_url) if "page=" in area_url else
#     #             f"{area_url}{'&' if '?' in area_url else '?'}page={page_num}"
#     #         )
            
#     #         print(f"\n--- Processing Page {page_num}/{total_pages} for {area_name} ---")
#     #         current_progress["current_page"] = page_num
#     #         scraped_current_progress["current_page"] = page_num
#     #         self.save_current_progress()
#     #         self.save_scraped_progress()
#     #         self.commit_progress(f"Started page {page_num} in {area_name}")
            
#     #         max_retries = 3
#     #         for attempt in range(max_retries):
#     #             try:
#     #                 restaurants_on_page = await self.get_page_restaurants(page_url, page_num)
#     #                 if not restaurants_on_page:
#     #                     raise Exception("No restaurants found")
#     #                 print(f"Found {len(restaurants_on_page)} restaurants on page {page_num}")
#     #                 break
#     #             except Exception as e:
#     #                 print(f"Error on page {page_num}: {e}")
#     #                 if attempt < max_retries - 1:
#     #                     print(f"Retrying ({attempt + 1}/{max_retries})...")
#     #                     await asyncio.sleep(5)
#     #                 else:
#     #                     print(f"Skipping page {page_num} after {max_retries} attempts")
#     #                     restaurants_on_page = []
            
#     #         if current_progress["total_restaurants"] == 0 or page_num > start_page:
#     #             current_progress["total_restaurants"] = len(restaurants_on_page)
#     #             scraped_current_progress["total_restaurants"] = len(restaurants_on_page)
#     #             if not is_resuming or page_num > start_page:
#     #                 current_progress["current_restaurant"] = 0
#     #                 scraped_current_progress["current_restaurant"] = 0
            
#     #         for rest_idx, restaurant in enumerate(restaurants_on_page):
#     #             rest_num = rest_idx + 1  # Start counting from 1
#     #             if rest_num <= current_progress["current_restaurant"]:
#     #                 print(f"Skipping processed restaurant {rest_num}/{len(restaurants_on_page)}: {restaurant['name']}")
#     #                 continue
                
#     #             # Check if restaurant is already in processed_restaurants before processing
#     #             if restaurant["name"] in current_progress["processed_restaurants"]:
#     #                 print(f"Skipping restaurant {rest_num}/{len(restaurants_on_page)}: {restaurant['name']} - Already processed previously")
#     #                 current_progress["current_restaurant"] = rest_num
#     #                 scraped_current_progress["current_restaurant"] = rest_num
#     #                 self.save_current_progress()
#     #                 self.save_scraped_progress()
#     #                 self.print_progress_details()
#     #                 self.commit_progress(f"Skipped restaurant {restaurant['name']} on page {page_num} in {area_name} (already processed)")
#     #                 continue
                
#     #             current_progress["current_restaurant"] = rest_num
#     #             scraped_current_progress["current_restaurant"] = rest_num
                
#     #             if any(category in restaurant['cuisine'] for category in skip_categories):
#     #                 print(f"\nSkipping restaurant {rest_num}/{len(restaurants_on_page)}: {restaurant['name']} - Category: {restaurant['cuisine']}")
#     #                 if restaurant["name"] not in current_progress["processed_restaurants"]:
#     #                     current_progress["processed_restaurants"].append(restaurant["name"])
#     #                     scraped_current_progress["processed_restaurants"].append(restaurant["name"])
#     #                 self.save_current_progress()
#     #                 self.save_scraped_progress()
#     #                 self.print_progress_details()
#     #                 self.commit_progress(f"Skipped restaurant {restaurant['name']} on page {page_num} in {area_name}")
#     #                 continue
                
#     #             print(f"\nProcessing restaurant {rest_num}/{len(restaurants_on_page)} on page {page_num}: {restaurant['name']}")
                
#     #             try:
#     #                 restaurant.setdefault("menu_items", {})
#     #                 restaurant.setdefault("info", {})
#     #                 restaurant.setdefault("reviews", {})
                    
#     #                 menu_data = await self.talabat_scraper.get_restaurant_menu(restaurant['url'])
#     #                 if menu_data:
#     #                     restaurant['menu_items'] = menu_data
                    
#     #                 info_data = await self.talabat_scraper.get_restaurant_info(restaurant['url'])
#     #                 if info_data:
#     #                     restaurant['info'] = info_data
                    
#     #                 if restaurant['info'].get('Reviews URL') and restaurant['info']['Reviews URL'] != 'Not Available':
#     #                     reviews_data = self.talabat_scraper.get_reviews_data(restaurant['info']['Reviews URL'])
#     #                     if reviews_data:
#     #                         restaurant['reviews'] = reviews_data
                    
#     #                 all_area_results.append(restaurant)
#     #                 if restaurant["name"] not in current_progress["processed_restaurants"]:
#     #                     current_progress["processed_restaurants"].append(restaurant["name"])
#     #                     scraped_current_progress["processed_restaurants"].append(restaurant["name"])
#     #                 self.scraped_progress["all_results"][area_name] = all_area_results
#     #                 self.save_current_progress()
#     #                 self.save_scraped_progress()
#     #                 self.print_progress_details()
#     #                 self.commit_progress(f"Processed restaurant {restaurant['name']} on page {page_num} in {area_name}")
#     #                 await asyncio.sleep(2)
                
#     #             except Exception as e:
#     #                 print(f"Error processing restaurant {rest_num}/{len(restaurants_on_page)}: {restaurant['name']}: {str(e)}")
#     #                 import traceback
#     #                 traceback.print_exc()
#     #                 if restaurant["name"] not in current_progress["processed_restaurants"]:
#     #                     current_progress["processed_restaurants"].append(restaurant["name"])
#     #                     scraped_current_progress["processed_restaurants"].append(restaurant["name"])
#     #                 self.save_current_progress()
#     #                 self.save_scraped_progress()
#     #                 self.print_progress_details()
#     #                 self.commit_progress(f"Error processing restaurant {restaurant['name']} on page {page_num} in {area_name}")
            
#     #         current_progress["completed_pages"].append(page_num)
#     #         scraped_current_progress["completed_pages"].append(page_num)
#     #         current_progress["current_restaurant"] = 0
#     #         scraped_current_progress["current_restaurant"] = 0
#     #         self.save_current_progress()
#     #         self.save_scraped_progress()
#     #         self.print_progress_details()
#     #         self.commit_progress(f"Completed page {page_num} in {area_name}")
#     #         await asyncio.sleep(3)
        
#     #     json_filename = os.path.join(self.output_dir, f"{area_name}.json")
#     #     with open(json_filename, 'w', encoding='utf-8') as f:
#     #         json.dump(all_area_results, f, indent=2, ensure_ascii=False)
        
#     #     workbook = Workbook()
#     #     self.create_excel_sheet(workbook, area_name, all_area_results)
#     #     excel_filename = os.path.join(self.output_dir, f"{area_name}.xlsx")
#     #     workbook.save(excel_filename)
#     #     print(f"Excel file saved: {excel_filename}")
        
#     #     if self.upload_to_drive(excel_filename):
#     #         print(f"Uploaded {excel_filename} to Google Drive")
#     #     else:
#     #         print(f"Failed to upload {excel_filename} to Google Drive")
        
#     #     current_progress.update({
#     #         "area_name": None,
#     #         "current_page": 0,
#     #         "total_pages": 0,
#     #         "current_restaurant": 0,
#     #         "total_restaurants": 0,
#     #         "processed_restaurants": [],
#     #         "completed_pages": []
#     #     })
#     #     scraped_current_progress.update(current_progress)
#     #     self.save_current_progress()
#     #     self.save_scraped_progress()
#     #     self.print_progress_details()
#     #     self.commit_progress(f"Completed area {area_name}")
        
#     #     print(f"Saved {len(all_area_results)} restaurants for {area_name}")
#     #     return all_area_results

#     async def scrape_and_save_area(self, area_name: str, area_url: str) -> List[Dict]:
#         print(f"\n{'='*50}")
#         print(f"SCRAPING AREA: {area_name}")
#         print(f"URL: {area_url}")
#         print(f"{'='*50}\n")
        
#         all_area_results = self.scraped_progress["all_results"].get(area_name, [])
#         current_progress = self.current_progress["current_progress"]
#         scraped_current_progress = self.scraped_progress["current_progress"]
        
#         is_resuming = current_progress["area_name"] == area_name
#         start_page = current_progress["current_page"] if is_resuming else 1
#         start_restaurant = current_progress["current_restaurant"] if is_resuming else 0
        
#         if is_resuming:
#             print(f"Resuming area {area_name} from page {start_page} restaurant {start_restaurant + 1 if start_restaurant > 0 else 1}")
#         else:
#             current_progress.update({
#                 "area_name": area_name,
#                 "current_page": start_page,
#                 "total_pages": 0,
#                 "current_restaurant": 0,
#                 "total_restaurants": 0,
#                 "processed_restaurants": [],
#                 "completed_pages": []
#             })
#             scraped_current_progress.update(current_progress)
#             self.save_current_progress()
#             self.save_scraped_progress()
#             self.commit_progress(f"Started scraping area {area_name}")
        
#         skip_categories = {"Grocery, Convenience Store", "Pharmacy", "Flowers", "Electronics", "Grocery, Hypermarket"}
        
#         if current_progress["total_pages"] == 0:
#             total_pages = await self.determine_total_pages(area_url)
#             current_progress["total_pages"] = total_pages
#             scraped_current_progress["total_pages"] = total_pages
#             self.save_current_progress()
#             self.save_scraped_progress()
#             self.commit_progress(f"Determined {total_pages} pages for {area_name}")
#         else:
#             total_pages = current_progress["total_pages"]
        
#         print(f"Total pages for {area_name}: {total_pages}")
        
#         for page_num in range(start_page, total_pages + 1):
#             if page_num in current_progress["completed_pages"]:
#                 print(f"Skipping completed page {page_num}")
#                 continue
            
#             page_url = area_url if page_num == 1 else (
#                 re.sub(r'page=\d+', f'page={page_num}', area_url) if "page=" in area_url else
#                 f"{area_url}{'&' if '?' in area_url else '?'}page={page_num}"
#             )
            
#             print(f"\n--- Processing Page {page_num}/{total_pages} for {area_name} ---")
#             current_progress["current_page"] = page_num
#             scraped_current_progress["current_page"] = page_num
#             self.save_current_progress()
#             self.save_scraped_progress()
#             self.commit_progress(f"Started page {page_num} in {area_name}")
            
#             max_retries = 3
#             for attempt in range(max_retries):
#                 try:
#                     restaurants_on_page = await self.get_page_restaurants(page_url, page_num)
#                     if not restaurants_on_page:
#                         raise Exception("No restaurants found")
#                     print(f"Found {len(restaurants_on_page)} restaurants on page {page_num}")
#                     break
#                 except Exception as e:
#                     print(f"Error on page {page_num}: {e}")
#                     logging.error(f"Error on page {page_num}: {e}")
#                     if attempt < max_retries - 1:
#                         print(f"Retrying ({attempt + 1}/{max_retries})...")
#                         await asyncio.sleep(5)
#                     else:
#                         print(f"Skipping page {page_num} after {max_retries} attempts")
#                         restaurants_on_page = []
            
#             if current_progress["total_restaurants"] == 0 or page_num > start_page:
#                 current_progress["total_restaurants"] = len(restaurants_on_page)
#                 scraped_current_progress["total_restaurants"] = len(restaurants_on_page)
#                 if not is_resuming or page_num > start_page:
#                     current_progress["current_restaurant"] = 0
#                     scraped_current_progress["current_restaurant"] = 0
            
#             for rest_idx, restaurant in enumerate(restaurants_on_page):
#                 rest_num = rest_idx + 1
#                 if rest_num <= current_progress["current_restaurant"]:
#                     print(f"Skipping processed restaurant {rest_num}/{len(restaurants_on_page)}: {restaurant['name']}")
#                     continue
                
#                 if restaurant["name"] in current_progress["processed_restaurants"]:
#                     print(f"Skipping restaurant {rest_num}/{len(restaurants_on_page)}: {restaurant['name']} - Already processed previously")
#                     current_progress["current_restaurant"] = rest_num
#                     scraped_current_progress["current_restaurant"] = rest_num
#                     self.save_current_progress()
#                     self.save_scraped_progress()
#                     self.print_progress_details()
#                     self.commit_progress(f"Skipped restaurant {restaurant['name']} on page {page_num} in {area_name} (already processed)")
#                     continue
                
#                 current_progress["current_restaurant"] = rest_num
#                 scraped_current_progress["current_restaurant"] = rest_num
                
#                 if any(category in restaurant['cuisine'] for category in skip_categories):
#                     print(f"\nSkipping restaurant {rest_num}/{len(restaurants_on_page)}: {restaurant['name']} - Category: {restaurant['cuisine']}")
#                     if restaurant["name"] not in current_progress["processed_restaurants"]:
#                         current_progress["processed_restaurants"].append(restaurant["name"])
#                         scraped_current_progress["processed_restaurants"].append(restaurant["name"])
#                     self.save_current_progress()
#                     self.save_scraped_progress()
#                     self.print_progress_details()
#                     self.commit_progress(f"Skipped restaurant {restaurant['name']} on page {page_num} in {area_name}")
#                     continue
                
#                 print(f"\nProcessing restaurant {rest_num}/{len(restaurants_on_page)} on page {page_num}: {restaurant['name']}")
                
#                 try:
#                     restaurant.setdefault("menu_items", {})
#                     restaurant.setdefault("info", {})
#                     restaurant.setdefault("reviews", {})
                    
#                     menu_data = await self.talabat_scraper.get_restaurant_menu(restaurant['url'])
#                     if menu_data:
#                         restaurant['menu_items'] = menu_data
                    
#                     info_data = await self.talabat_scraper.get_restaurant_info(restaurant['url'])
#                     if info_data:
#                         restaurant['info'] = info_data
                    
#                     if restaurant['info'].get('Reviews URL') and restaurant['info']['Reviews URL'] != 'Not Available':
#                         reviews_data = self.talabat_scraper.get_reviews_data(restaurant['info']['Reviews URL'])
#                         if reviews_data:
#                             restaurant['reviews'] = reviews_data
                    
#                     all_area_results.append(restaurant)
#                     if restaurant["name"] not in current_progress["processed_restaurants"]:
#                         current_progress["processed_restaurants"].append(restaurant["name"])
#                         scraped_current_progress["processed_restaurants"].append(restaurant["name"])
#                     self.scraped_progress["all_results"][area_name] = all_area_results
#                     self.save_current_progress()
#                     self.save_scraped_progress()
#                     self.print_progress_details()
#                     self.commit_progress(f"Processed restaurant {restaurant['name']} on page {page_num} in {area_name}")
#                     await asyncio.sleep(2)
                
#                 except Exception as e:
#                     print(f"Error processing restaurant {rest_num}/{len(restaurants_on_page)}: {restaurant['name']}: {e}")
#                     logging.error(f"Error processing restaurant {restaurant['name']}: {e}")
#                     import traceback
#                     traceback.print_exc()
#                     if restaurant["name"] not in current_progress["processed_restaurants"]:
#                         current_progress["processed_restaurants"].append(restaurant["name"])
#                         scraped_current_progress["processed_restaurants"].append(restaurant["name"])
#                     self.save_current_progress()
#                     self.save_scraped_progress()
#                     self.print_progress_details()
#                     self.commit_progress(f"Error processing restaurant {restaurant['name']} on page {page_num} in {area_name}")
            
#             if page_num not in current_progress["completed_pages"]:
#                 current_progress["completed_pages"].append(page_num)
#                 scraped_current_progress["completed_pages"].append(page_num)
#             current_progress["current_restaurant"] = 0
#             scraped_current_progress["current_restaurant"] = 0
#             self.save_current_progress()
#             self.save_scraped_progress()
#             self.print_progress_details()
#             self.commit_progress(f"Completed page {page_num} in {area_name}")
#             await asyncio.sleep(3)
        
#         json_filename = os.path.join(self.output_dir, f"{area_name}.json")
#         with open(json_filename, 'w', encoding='utf-8') as f:
#             json.dump(all_area_results, f, indent=2, ensure_ascii=False)
        
#         workbook = Workbook()
#         self.create_excel_sheet(workbook, area_name, all_area_results)
#         excel_filename = os.path.join(self.output_dir, f"{area_name}.xlsx")
#         workbook.save(excel_filename)
#         print(f"Excel file saved: {excel_filename}")
        
#         if self.upload_to_drive(excel_filename):
#             print(f"Uploaded {excel_filename} to Google Drive")
#         else:
#             print(f"Failed to upload {excel_filename} to Google Drive")
        
#         current_progress.update({
#             "area_name": None,
#             "current_page": 0,
#             "total_pages": 0,
#             "current_restaurant": 0,
#             "total_restaurants": 0,
#             "processed_restaurants": [],
#             "completed_pages": []
#         })
#         scraped_current_progress.update(current_progress)
#         self.save_current_progress()
#         self.save_scraped_progress()
#         self.print_progress_details()
#         self.commit_progress(f"Completed area {area_name}")
        
#         print(f"Saved {len(all_area_results)} restaurants for {area_name}")
#         return all_area_results

#     async def determine_total_pages(self, area_url: str) -> int:
#         print(f"Determining total pages for URL: {area_url}")
#         try:
#             async with async_playwright() as p:
#                 browser = await p.firefox.launch(headless=True)
#                 context = await browser.new_context(
#                     viewport={'width': 1920, 'height': 1080},
#                     user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
#                 )
#                 page = await context.new_page()
#                 page.set_default_timeout(120000)
                
#                 response = await page.goto(area_url, wait_until='domcontentloaded')
#                 if not response or not response.ok:
#                     print(f"Failed to load page: {response.status if response else 'No response'}")
#                     return 1
                
#                 await page.wait_for_selector("ul[data-test='pagination'], .vendor-card, [data-testid='restaurant-a']", timeout=30000)
                
#                 last_page = 1
#                 pagination = await page.query_selector("ul[data-test='pagination']")
#                 if pagination:
#                     items = await pagination.query_selector_all("li[data-testid='paginate-link']")
#                     if items and len(items) > 1:
#                         last_page_item = items[-2]
#                         last_page_link = await last_page_item.query_selector("a[page]")
#                         if last_page_link:
#                             last_page_attr = await last_page_link.get_attribute("page")
#                             if last_page_attr and last_page_attr.isdigit():
#                                 last_page = int(last_page_attr)
                
#                 await browser.close()
#                 return last_page
#         except Exception as e:
#             print(f"Error determining total pages: {e}")
#             return 1

#     async def get_page_restaurants(self, page_url: str, page_num: int) -> List[Dict]:
#         browser = None
#         try:
#             async with async_playwright() as p:
#                 browser = await p.firefox.launch(headless=True)
#                 context = await browser.new_context(
#                     viewport={'width': 1920, 'height': 1080},
#                     user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
#                 )
#                 page = await context.new_page()
#                 page.set_default_timeout(120000)
                
#                 response = await page.goto(page_url, wait_until='domcontentloaded')
#                 if not response or not response.ok:
#                     print(f"Failed to load page {page_num}: {response.status if response else 'No response'}")
#                     return []
                
#                 await page.wait_for_selector(".vendor-card, [data-testid='restaurant-a']", timeout=30000)
#                 return await self.talabat_scraper._extract_restaurants_from_page(page, page_num)
#         except Exception as e:
#             print(f"Error getting page restaurants: {e}")
#             import traceback
#             traceback.print_exc()
#             return []
#         finally:
#             if browser:
#                 await browser.close()

#     def create_excel_sheet(self, workbook, sheet_name: str, data: List[Dict]):
#         sheet = workbook.create_sheet(title=sheet_name)
#         try:
#             simplified_data = []
#             for restaurant in data:
#                 restaurant_info = {
#                     "Name": restaurant.get("name", ""),
#                     "Cuisine": restaurant.get("cuisine", ""),
#                     "Rating": restaurant.get("rating", ""),
#                     "Delivery Time": restaurant.get("delivery_time", ""),
#                     "Delivery Fee": restaurant.get("delivery_fee", ""),
#                     "Min Order": restaurant.get("min_order", ""),
#                     "URL": restaurant.get("url", ""),
#                 }
#                 if restaurant.get("info"):
#                     restaurant_info.update({
#                         "Address": restaurant["info"].get("Address", ""),
#                         "Working Hours": restaurant["info"].get("Working Hours", ""),
#                     })
#                 if restaurant.get("reviews") and restaurant["reviews"].get("Rating_value"):
#                     restaurant_info.update({
#                         "Rating Value": restaurant["reviews"]["Rating_value"],
#                         "Ratings Count": restaurant["reviews"].get("Ratings_count", ""),
#                         "Reviews Count": restaurant["reviews"].get("Reviews_count", ""),
#                     })
#                 if restaurant.get("menu_items"):
#                     restaurant_info["Menu Categories"] = len(restaurant["menu_items"])
#                     item_count = sum(len(items) for items in restaurant["menu_items"].values())
#                     restaurant_info["Menu Items"] = item_count
#                 simplified_data.append(restaurant_info)
            
#             if simplified_data:
#                 df = pd.DataFrame(simplified_data)
#                 for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
#                     for c_idx, value in enumerate(row, 1):
#                         sheet.cell(row=r_idx, column=c_idx, value=value)
#                 for column in sheet.columns:
#                     max_length = max(len(str(cell.value or "")) for cell in column)
#                     column_letter = get_column_letter(column[0].column)
#                     sheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
#             else:
#                 sheet.cell(row=1, column=1, value="No data found for this area")
#         except Exception as e:
#             print(f"Error creating Excel sheet for {sheet_name}: {str(e)}")
#             sheet.cell(row=1, column=1, value=f"Error processing data: {str(e)}")
        
#     # def upload_to_drive(self, file_path):
#     #     print(f"\nUploading {file_path} to Google Drive...")
#     #     try:
#     #         if not self.drive_uploader.authenticate():
#     #             print("Failed to authenticate with Google Drive")
#     #             return False
#     #         file_ids = self.drive_uploader.upload_to_multiple_folders(file_path)
#     #         return len(file_ids) == 2
#     #     except Exception as e:
#     #         print(f"Error uploading to Google Drive: {str(e)}")
#     #         return False
    
#     @retry(tries=3, delay=2, backoff=2)
#     def upload_to_drive(self, file_path):
#         print(f"\nUploading {file_path} to Google Drive...")
#         try:
#             credentials_json = os.environ.get('TALABAT_GCLOUD_KEY_JSON')
#             if not credentials_json:
#                 print("Error: TALABAT_GCLOUD_KEY_JSON environment variable is empty or not set!")
#                 return False
#             self.drive_uploader = SavingOnDrive(credentials_json=credentials_json)
#             if not self.drive_uploader.authenticate():
#                 print("Failed to authenticate with Google Drive. Check TALABAT_GCLOUD_KEY_JSON validity.")
#                 return False
#             file_ids = self.drive_uploader.upload_to_multiple_folders(file_path)
#             success = len(file_ids) == 2
#             if success:
#                 print(f"Successfully uploaded {file_path} to Google Drive")
#             else:
#                 print(f"Failed to upload {file_path}: Incomplete upload to folders")
#             return success
#         except Exception as e:
#             print(f"Error uploading to Google Drive: {str(e)}")
#             return False

#     async def run(self):
#         ahmadi_areas = [
#             ("الظهر", "https://www.talabat.com/kuwait/restaurants/59/dhaher"),
#             ("الرقه", "https://www.talabat.com/kuwait/restaurants/37/riqqa"),
#             ("هدية", "https://www.talabat.com/kuwait/restaurants/30/hadiya"),
#             ("المنقف", "https://www.talabat.com/kuwait/restaurants/32/mangaf"),
#             ("أبو حليفة", "https://www.talabat.com/kuwait/restaurants/2/abu-halifa"),
#             ("الفنطاس", "https://www.talabat.com/kuwait/restaurants/38/fintas"),
#             ("العقيلة", "https://www.talabat.com/kuwait/restaurants/79/egaila"),
#             ("الصباحية", "https://www.talabat.com/kuwait/restaurants/31/sabahiya"),
#             ("الأحمدي", "https://www.talabat.com/kuwait/restaurants/3/al-ahmadi"),
#             ("الفحيحيل", "https://www.talabat.com/kuwait/restaurants/5/fahaheel"),
#             ("شرق الأحمدي", "https://www.talabat.com/kuwait/restaurants/3/al-ahmadi"),
#             ("ضاحية علي صباح السالم", "https://www.talabat.com/kuwait/restaurants/82/ali-sabah-al-salem-umm-al-hayman"),
#             ("ميناء عبد الله", "https://www.talabat.com/kuwait/restaurants/100/mina-abdullah"),
#             ("بنيدر", "https://www.talabat.com/kuwait/restaurants/6650/bnaider"),
#             ("الزور", "https://www.talabat.com/kuwait/restaurants/2053/zour"),
#             ("الجليعة", "https://www.talabat.com/kuwait/restaurants/6860/al-julaiaa"),
#             ("المهبولة", "https://www.talabat.com/kuwait/restaurants/24/mahboula"),
#             ("النويصيب", "https://www.talabat.com/kuwait/restaurants/2054/nuwaiseeb"),
#             ("الخيران", "https://www.talabat.com/kuwait/restaurants/2726/khairan"),
#             ("الوفرة", "https://www.talabat.com/kuwait/restaurants/2057/wafra-farms"),
#             ("ضاحية فهد الأحمد", "https://www.talabat.com/kuwait/restaurants/98/fahad-al-ahmed"),
#             ("ضاحية جابر العلي", "https://www.talabat.com/kuwait/restaurants/60/jaber-al-ali"),
#             ("مدينة صباح الأحمد السكنية", "https://www.talabat.com/kuwait/restaurants/6931/sabah-al-ahmad-2"),
#             ("مدينة صباح الأحمد البحرية", "https://www.talabat.com/kuwait/restaurants/2726/khairan"),
#             ("ميناء الأحمدي", "https://www.talabat.com/kuwait/restaurants/3/al-ahmadi")
#         ]
        
#         excel_filename = os.path.join(self.output_dir, "الاحمدي.xlsx")
#         workbook = Workbook()
#         if "Sheet" in workbook.sheetnames:
#             workbook.remove(workbook["Sheet"])
        
#         completed_areas = self.current_progress["completed_areas"]
#         current_area_index = self.current_progress["current_area_index"]
        
#         print(f"Starting from area index {current_area_index}")
#         print(f"Already completed areas: {', '.join(completed_areas) if completed_areas else 'None'}")
        
#         resuming_area = self.current_progress["current_progress"]["area_name"]
#         if resuming_area:
#             for idx, (area_name, _) in enumerate(ahmadi_areas):
#                 if area_name == resuming_area:
#                     print(f"Resuming from area {resuming_area} (index {idx})")
#                     current_area_index = idx
#                     self.current_progress["current_area_index"] = idx
#                     self.scraped_progress["current_area_index"] = idx
#                     self.save_current_progress()
#                     self.save_scraped_progress()
#                     self.commit_progress(f"Resuming from area {resuming_area}")
#                     break
        
#         for idx, (area_name, area_url) in enumerate(ahmadi_areas):
#             if area_name in completed_areas and area_name != resuming_area:
#                 print(f"Skipping completed area: {area_name}")
#                 continue
#             if idx < current_area_index:
#                 print(f"Skipping area {area_name} (index {idx} < {current_area_index})")
#                 continue
            
#             self.current_progress["current_area_index"] = idx
#             self.scraped_progress["current_area_index"] = idx
#             self.save_current_progress()
#             self.save_scraped_progress()
#             self.commit_progress(f"Starting area {area_name} at index {idx}")
            
#             try:
#                 area_results = await self.scrape_and_save_area(area_name, area_url)
#                 self.create_excel_sheet(workbook, area_name, area_results)
#                 workbook.save(excel_filename)
#                 print(f"Updated Excel file: {excel_filename}")
                
#                 if area_name not in completed_areas:
#                     completed_areas.append(area_name)
#                     self.current_progress["completed_areas"] = completed_areas
#                     self.scraped_progress["completed_areas"] = completed_areas
#                 self.save_current_progress()
#                 self.save_scraped_progress()
#                 self.print_progress_details()
#                 self.commit_progress(f"Completed area {area_name} in run")
#                 await asyncio.sleep(5)
            
#             except Exception as e:
#                 print(f"Error processing area {area_name}: {e}")
#                 logging.error(f"Error processing area {area_name}: {e}")
#                 import traceback
#                 traceback.print_exc()
#                 self.save_current_progress()
#                 self.save_scraped_progress()
#                 self.commit_progress(f"Progress update after error in {area_name}")
        
#         workbook.save(excel_filename)
#         combined_json_filename = os.path.join(self.output_dir, "الاحمدي_all.json")
#         with open(combined_json_filename, 'w', encoding='utf-8') as f:
#             json.dump(self.scraped_progress["all_results"], f, indent=2, ensure_ascii=False)
        
#         print(f"\n{'='*50}")
#         print(f"SCRAPING COMPLETED")
#         print(f"Excel file saved: {excel_filename}")
#         print(f"Combined JSON saved: {combined_json_filename}")
        
#         if len(completed_areas) == len(ahmadi_areas):
#             if self.upload_to_drive(excel_filename):
#                 print(f"Uploaded Excel file to Google Drive")
#             else:
#                 print(f"Failed to upload Excel file to Google Drive")
#         else:
#             print(f"Scraping incomplete ({len(completed_areas)}/{len(ahmadi_areas)} areas)")
        
#         self.commit_progress("Final progress update after run")

# # def create_credentials_file():
# #     try:
# #         credentials_json = os.environ.get('TALABAT_GCLOUD_KEY_JSON')
# #         if not credentials_json:
# #             print("ERROR: TALABAT_GCLOUD_KEY_JSON not found!")
# #             return False
# #         with open('credentials.json', 'w') as f:
# #             f.write(credentials_json)
# #         print("Created credentials.json")
# #         return True
# #     except Exception as e:
# #         print(f"ERROR: Failed to create credentials.json: {str(e)}")
# #         return False

# # async def main():
# #     if not create_credentials_file():
# #         print("Could not create credentials.json")
# #         sys.exit(1)
    
# #     if not os.path.exists('credentials.json'):
# #         print("ERROR: credentials.json not found!")
# #         sys.exit(1)
    
# #     try:
# #         scraper = MainScraper()
# #         await scraper.run()
# #     except KeyboardInterrupt:
# #         print("\nInterrupted. Saving progress...")
# #         if 'scraper' in locals():
# #             scraper.save_current_progress()
# #             scraper.save_scraped_progress()
# #             scraper.commit_progress("Progress saved after interruption")
# #         print("Progress saved. Exiting.")
# #     except Exception as e:
# #         print(f"Critical error: {e}")
# #         import traceback
# #         traceback.print_exc()
# #         if 'scraper' in locals():
# #             scraper.save_current_progress()
# #             scraper.save_scraped_progress()
# #             scraper.commit_progress("Progress saved after critical error")
# #         sys.exit(1)
# async def main():
#     try:
#         scraper = MainScraper()
#         await scraper.run()
#     except KeyboardInterrupt:
#         print("\nInterrupted. Saving progress...")
#         if 'scraper' in locals():
#             scraper.save_current_progress()
#             scraper.save_scraped_progress()
#             scraper.commit_progress("Progress saved after interruption")
#         print("Progress saved. Exiting.")
#     except Exception as e:
#         print(f"Critical error: {e}")
#         import traceback
#         traceback.print_exc()
#         if 'scraper' in locals():
#             scraper.save_current_progress()
#             scraper.save_scraped_progress()
#             scraper.commit_progress("Progress saved after critical error")
#         sys.exit(1)
        
# if __name__ == "__main__":
#     scraper = MainScraper()
#     scraper.print_progress_details()
#     asyncio.run(scraper.run())
